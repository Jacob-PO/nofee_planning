#!/usr/bin/env python3
"""
이미지에서 텍스트 색상을 정확하게 추출하는 스크립트
OCR 결과의 텍스트 위치 정보를 활용하여 해당 위치의 글자색을 추출
"""

import cv2
import numpy as np
from pathlib import Path
import json
from collections import Counter
from typing import Dict, Tuple, List
import openpyxl
from openpyxl.styles import Font
import sys


def get_dominant_text_color(img: np.ndarray, x1: int, y1: int, x2: int, y2: int) -> str:
    """
    주어진 영역에서 텍스트 색상을 추출
    """
    # 영역 추출
    region = img[y1:y2, x1:x2]
    if region.size == 0:
        return "black"
    
    # 이미지를 그레이스케일로 변환하여 텍스트 위치 찾기
    gray = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
    
    # 텍스트는 보통 배경보다 어두움 - 어두운 픽셀 찾기
    # Otsu's thresholding으로 텍스트 영역 찾기
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    
    # 텍스트 픽셀 위치 찾기
    text_pixels_coords = np.where(binary == 255)
    
    if len(text_pixels_coords[0]) == 0:
        return "black"
    
    # 텍스트 픽셀의 색상 수집
    text_colors = []
    for i in range(len(text_pixels_coords[0])):
        y = text_pixels_coords[0][i]
        x = text_pixels_coords[1][i]
        # BGR to RGB
        color = region[y, x][::-1]
        text_colors.append(tuple(color))
    
    if not text_colors:
        return "black"
    
    # 색상 분석
    red_count = 0
    black_count = 0
    
    for color in text_colors:
        r, g, b = int(color[0]), int(color[1]), int(color[2])
        
        # 빨간색 판별: R=255이고 G, B가 낮은 경우 (실제 측정값 기반)
        if r >= 250 and g < 170 and b < 220:
            red_count += 1
        # 약간 어두운 빨간색 (안전 마진)
        elif r >= 200 and g < 100 and b < 100 and r > g * 2 and r > b * 2:
            red_count += 1
        else:
            black_count += 1
    
    # 빨간색 픽셀이 전체의 30% 이상이면 빨간색으로 판단
    if red_count > len(text_colors) * 0.3:
        return "red"
    else:
        return "black"


def extract_text_colors_from_ocr(image_path: Path, ocr_result: Dict) -> Dict[str, str]:
    """
    OCR 결과를 바탕으로 각 셀의 텍스트 색상 추출
    """
    # 이미지 읽기
    img = cv2.imread(str(image_path))
    if img is None:
        print(f"이미지를 읽을 수 없습니다: {image_path}")
        return {}
    
    text_colors = {}
    
    # OCR 결과에서 텍스트 위치 정보 추출
    if 'images' in ocr_result:
        for image in ocr_result['images']:
            if 'tables' in image:
                for table in image['tables']:
                    if 'cells' in table:
                        for cell in table['cells']:
                            row_idx = cell.get('rowIndex', 0)
                            col_idx = cell.get('columnIndex', 0)
                            cell_key = f"{row_idx}_{col_idx}"
                            
                            # 셀 내의 텍스트 라인들 확인
                            if 'cellTextLines' in cell and cell['cellTextLines']:
                                cell_has_red = False
                                
                                for text_line in cell['cellTextLines']:
                                    if 'boundingPoly' in text_line and 'vertices' in text_line['boundingPoly']:
                                        vertices = text_line['boundingPoly']['vertices']
                                        if len(vertices) >= 4:
                                            x1 = int(min(v['x'] for v in vertices))
                                            y1 = int(min(v['y'] for v in vertices))
                                            x2 = int(max(v['x'] for v in vertices))
                                            y2 = int(max(v['y'] for v in vertices))
                                            
                                            # 영역이 이미지 범위 내에 있는지 확인
                                            x1 = max(0, x1)
                                            y1 = max(0, y1)
                                            x2 = min(img.shape[1], x2)
                                            y2 = min(img.shape[0], y2)
                                            
                                            if x2 > x1 and y2 > y1:
                                                color = get_dominant_text_color(img, x1, y1, x2, y2)
                                                if color == "red":
                                                    cell_has_red = True
                                
                                # 셀에 빨간색 텍스트가 하나라도 있으면 빨간색으로 표시
                                if cell_has_red:
                                    text_colors[cell_key] = "red"
                                else:
                                    text_colors[cell_key] = "black"
    
    return text_colors


def apply_text_colors_to_excel(excel_path: Path, text_colors: Dict[str, str], sheet_name: str):
    """
    Excel 파일에 텍스트 색상 적용
    """
    wb = openpyxl.load_workbook(excel_path)
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 테이블 시작 위치 찾기 - OCR row 0은 Excel row 3에 해당
        table_start_row = 3  # OCR row 0 = Excel row 3
        
        # 색상 적용
        for cell_key, color in text_colors.items():
            row_idx, col_idx = map(int, cell_key.split('_'))
            excel_row = table_start_row + row_idx
            excel_col = col_idx + 1
            
            cell = ws.cell(row=excel_row, column=excel_col)
            if cell.value:  # 값이 있는 셀만
                if color == "red":
                    cell.font = Font(color="FF0000", size=10)
                else:
                    cell.font = Font(color="000000", size=10)
        
        # 저장
        wb.save(excel_path)
        print(f"✅ {excel_path} 파일에 글자색 적용 완료")
    else:
        print(f"⚠️  {sheet_name} 시트를 찾을 수 없습니다")


def main():
    """메인 함수"""
    if len(sys.argv) < 2:
        print("사용법: python extract_text_colors.py <image_file>")
        sys.exit(1)
    
    image_path = Path(sys.argv[1])
    if not image_path.exists():
        print(f"이미지 파일을 찾을 수 없습니다: {image_path}")
        sys.exit(1)
    
    # JSON 파일 찾기 - OCR_결과 폴더에서
    json_pattern = image_path.stem + ".json"
    ocr_results_folder = Path(__file__).parent / "OCR" / "OCR_결과"
    json_files = list(ocr_results_folder.rglob(f"*/{json_pattern}"))
    
    if not json_files:
        print(f"OCR 결과 JSON 파일을 찾을 수 없습니다: {json_pattern}")
        sys.exit(1)
    
    json_path = json_files[0]
    print(f"📄 JSON 파일 사용: {json_path}")
    
    # OCR 결과 읽기
    with open(json_path, 'r', encoding='utf-8') as f:
        ocr_data = json.load(f)
    
    # 이미지 이름으로 결과 찾기
    ocr_result = ocr_data.get(image_path.name)
    if not ocr_result:
        # 첫 번째 키의 값 사용 (단일 파일 결과인 경우)
        if len(ocr_data) == 1:
            ocr_result = list(ocr_data.values())[0]
    
    if not ocr_result:
        print("OCR 결과를 찾을 수 없습니다")
        sys.exit(1)
    
    # 텍스트 색상 추출
    print(f"🎨 {image_path.name}에서 텍스트 색상 추출 중...")
    text_colors = extract_text_colors_from_ocr(image_path, ocr_result)
    
    # 결과 출력
    red_cells = sum(1 for c in text_colors.values() if c == "red")
    black_cells = sum(1 for c in text_colors.values() if c == "black")
    
    print(f"\n📊 색상 추출 결과:")
    print(f"  - 빨간색 텍스트: {red_cells}개 셀")
    print(f"  - 검정색 텍스트: {black_cells}개 셀")
    
    # Excel 파일 찾기 및 적용
    excel_pattern = image_path.stem + "_tables.xlsx"
    excel_files = list(ocr_results_folder.rglob(f"*/{excel_pattern}"))
    
    if excel_files:
        excel_path = excel_files[0]
        print(f"\n📑 Excel 파일에 색상 적용: {excel_path}")
        apply_text_colors_to_excel(excel_path, text_colors, image_path.name)
    else:
        print(f"\n⚠️  Excel 파일을 찾을 수 없습니다: {excel_pattern}")


if __name__ == "__main__":
    main()