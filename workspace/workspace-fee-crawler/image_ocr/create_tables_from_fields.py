#!/usr/bin/env python3
"""
OCR fields 데이터로부터 표 형태로 재구성하는 스크립트
General API에서 tables 데이터가 없을 때 fields를 이용해 표 생성
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import sys

def reconstruct_table_from_fields(json_file):
    """
    JSON 파일의 fields 데이터로부터 표를 재구성
    """
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 첫 번째 이미지 키
    img_key = list(data.keys())[0]
    print(f"처리 중: {img_key}")

    if 'images' not in data[img_key] or len(data[img_key]['images']) == 0:
        print("이미지 데이터가 없습니다.")
        return None

    img_data = data[img_key]['images'][0]

    if 'fields' not in img_data:
        print("fields 데이터가 없습니다.")
        return None

    fields = img_data['fields']
    print(f"전체 필드 수: {len(fields)}")

    # 좌표 기반으로 그리드 생성
    # Y 좌표로 행 그룹화, X 좌표로 열 정렬

    # 모든 필드의 좌표 수집
    field_positions = []
    for field in fields:
        if 'boundingPoly' not in field:
            continue

        vertices = field['boundingPoly']['vertices']
        if len(vertices) < 4:
            continue

        # 바운딩 박스의 중심점 계산
        x_coords = [v['x'] for v in vertices]
        y_coords = [v['y'] for v in vertices]

        center_x = sum(x_coords) / len(x_coords)
        center_y = sum(y_coords) / len(y_coords)

        min_y = min(y_coords)
        max_y = max(y_coords)
        min_x = min(x_coords)
        max_x = max(x_coords)

        field_positions.append({
            'text': field.get('inferText', ''),
            'confidence': field.get('inferConfidence', 0),
            'center_x': center_x,
            'center_y': center_y,
            'min_x': min_x,
            'max_x': max_x,
            'min_y': min_y,
            'max_y': max_y,
            'height': max_y - min_y
        })

    if not field_positions:
        print("좌표 정보가 없습니다.")
        return None

    # Y 좌표로 정렬
    field_positions.sort(key=lambda x: (x['center_y'], x['center_x']))

    # 행 그룹화 (Y 좌표가 비슷한 것들을 같은 행으로)
    rows = []
    current_row = []
    threshold = 15  # Y 좌표 차이 허용 범위

    for field in field_positions:
        if not current_row:
            current_row.append(field)
        else:
            # 이전 필드와 Y 좌표 비교
            prev_y = current_row[-1]['center_y']
            if abs(field['center_y'] - prev_y) < threshold:
                current_row.append(field)
            else:
                # 새로운 행 시작
                current_row.sort(key=lambda x: x['center_x'])
                rows.append(current_row)
                current_row = [field]

    if current_row:
        current_row.sort(key=lambda x: x['center_x'])
        rows.append(current_row)

    print(f"감지된 행 수: {len(rows)}")

    # Excel 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = img_key

    # 데이터 쓰기
    for row_idx, row_fields in enumerate(rows, 1):
        for col_idx, field in enumerate(row_fields, 1):
            ws.cell(row=row_idx, column=col_idx, value=field['text'])

    # 출력 파일명 생성
    json_path = Path(json_file)
    output_file = json_path.parent / f"{json_path.stem}_tables.xlsx"

    wb.save(output_file)
    print(f"✅ 표 파일 생성: {output_file}")

    return str(output_file)


def main():
    """메인 함수"""
    if len(sys.argv) > 1:
        json_file = sys.argv[1]
    else:
        # latest 폴더의 모든 JSON 파일 처리
        latest_dir = Path(__file__).parent / "output" / "latest"
        if not latest_dir.exists():
            # archive의 최신 폴더 찾기
            archive_dir = Path(__file__).parent / "output" / "archive"
            if archive_dir.exists():
                date_folders = sorted([f for f in archive_dir.iterdir() if f.is_dir()], reverse=True)
                if date_folders:
                    latest_dir = date_folders[0]

        json_files = list(latest_dir.glob("*.json"))
        if not json_files:
            print("처리할 JSON 파일이 없습니다.")
            return

        print(f"{len(json_files)}개의 JSON 파일을 찾았습니다.")

        for json_file in json_files:
            if "all_results" in json_file.name:
                continue  # 통합 파일 제외
            print(f"\n{'='*60}")
            reconstruct_table_from_fields(json_file)


if __name__ == "__main__":
    main()
