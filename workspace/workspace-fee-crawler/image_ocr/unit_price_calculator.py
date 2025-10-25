import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import re
from typing import Dict, List, Tuple
import os
import unicodedata
from pathlib import Path


class UnitPriceCalculator:
    """대리점별 단가 계산 로직을 처리하는 클래스"""
    
    def __init__(self):
        self.calculations_log = []
        
    def extract_number(self, text: str) -> float:
        """텍스트에서 숫자 추출"""
        if not text:
            return 0
        text = str(text).strip()
        # 순수 숫자인 경우만 처리 (음수 포함)
        if re.match(r'^-?\d+\.?\d*$', text):
            return float(text)
        return 0
    
    def get_rate_from_text(self, text: str) -> int:
        """텍스트에서 요금제 숫자 추출 (다양한 형식 지원)"""
        if not text:
            return 0
        
        # 다양한 요금제 패턴 매칭
        patterns = [
            r'(\d+)[kK]',           # 100K, 95k
            r'(\d+)군',             # 115군, 95군
            r'\((\d+)[kK]?\)',     # (115K), (95K)
            r'\b(\d{2,3})\b'       # 단독 2-3자리 숫자
        ]
        
        text_str = str(text)
        for pattern in patterns:
            match = re.search(pattern, text_str)
            if match:
                num = int(match.group(1))
                # 50-200 범위의 숫자만 요금제로 인식
                if 50 <= num <= 200:
                    return num
        
        return 0
    
    def match_product(self, text: str, keywords: List[str]) -> bool:
        """텍스트에 키워드가 포함되어 있는지 확인 (대소문자 무시)"""
        if not text:
            return False
        text_lower = str(text).lower()
        return any(keyword.lower() in text_lower for keyword in keywords)
    
    def is_pure_number(self, value) -> bool:
        """순수 숫자인지 확인"""
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return True
        if isinstance(value, str):
            value_str = value.strip()
            # 010은 신규가입을 의미하므로 제외
            if value_str == "010":
                return False
            return bool(re.match(r'^-?\d+\.?\d*$', value_str))
        return False
    
    def get_cell_background_color(self, cell) -> str:
        """셀의 배경색 가져오기"""
        if hasattr(cell, 'fill') and cell.fill and hasattr(cell.fill, 'start_color'):
            color = cell.fill.start_color
            if hasattr(color, 'rgb') and color.rgb:
                # RGB 값이 문자열인 경우
                if isinstance(color.rgb, str) and len(color.rgb) >= 6:
                    rgb = color.rgb[-6:]  # ARGB에서 RGB만 추출
                    r = int(rgb[0:2], 16)
                    g = int(rgb[2:4], 16)
                    b = int(rgb[4:6], 16)
                    
                    # 노란색 감지 (R,G 높고 B 낮음)
                    if r > 200 and g > 200 and b < 100:
                        return "yellow"
                    # 파란색 감지 (B 높고 R,G 낮음)
                    elif b > 200 and r < 100 and g < 100:
                        return "blue"
                    # 빨간색 감지 (R 높고 G,B 낮음)
                    elif r > 200 and g < 100 and b < 100:
                        return "red"
        
        return "none"
    
    def is_red_text(self, cell) -> bool:
        """텍스트 색상이 빨간색인지 확인"""
        if hasattr(cell, 'font') and cell.font and hasattr(cell.font, 'color'):
            color = cell.font.color
            if hasattr(color, 'rgb') and color.rgb:
                rgb_str = str(color.rgb)
                
                # ARGB 형식 (8자리) 처리
                if len(rgb_str) == 8:
                    # ARGB에서 RGB만 추출 (첫 2자리 제외)
                    rgb = rgb_str[2:]
                elif len(rgb_str) == 6:
                    # 이미 RGB 형식
                    rgb = rgb_str
                else:
                    return False
                    
                try:
                    r = int(rgb[0:2], 16)
                    g = int(rgb[2:4], 16)
                    b = int(rgb[4:6], 16)
                    
                    # 빨간색 감지 (정확한 빨간색: FF0000)
                    if r > 200 and g < 50 and b < 50:
                        return True
                except ValueError:
                    pass
        return False
    
    def calculate_kt_더블유(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """KT_더블유 계산 로직"""
        print("KT_더블유 계산 시작")
        
        for sheet_name in wb.sheetnames:
            if "_분석" in sheet_name:
                continue
                
            ws = wb[sheet_name]
            
            # 모든 셀에 대해 계산
            for row_idx, row in enumerate(ws.iter_rows(), 1):
                for col_idx, cell in enumerate(row, 1):
                    if cell.value and self.is_pure_number(cell.value):
                        original_value = self.extract_number(cell.value)
                        if original_value != 0:
                            # 최종 계산 (추가 없이 10000만 곱하기)
                            final_value = original_value * 10000
                            cell.value = final_value
                            self.calculations_log.append(f"{sheet_name} {cell.coordinate}: {original_value} * 10000 = {final_value}")
        
        return wb
    
    def calculate_kt_맥스(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """KT_맥스 계산 로직"""
        print("KT_맥스 계산 시작")
        
        for sheet_name in wb.sheetnames:
            if "_분석" in sheet_name:
                continue
                
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and self.is_pure_number(cell.value):
                        original_value = self.extract_number(cell.value)
                        if original_value != 0:
                            # 모든 상품에 +3
                            new_value = original_value + 3
                            self.calculations_log.append(f"{sheet_name} {cell.coordinate}: {original_value} + 3 = {new_value}")
                            
                            # 배경색에 따른 추가 계산
                            bg_color = self.get_cell_background_color(cell)
                            if bg_color == "yellow":
                                new_value += 30
                                self.calculations_log.append(f"{sheet_name} {cell.coordinate}: +30 (노란색) = {new_value}")
                            elif bg_color == "blue":
                                new_value += 20
                                self.calculations_log.append(f"{sheet_name} {cell.coordinate}: +20 (파란색) = {new_value}")
                            elif bg_color == "red":
                                new_value += 10
                                self.calculations_log.append(f"{sheet_name} {cell.coordinate}: +10 (빨간색) = {new_value}")
                            
                            # 최종 계산
                            final_value = new_value * 10000
                            cell.value = final_value
        
        return wb
    
    def calculate_lg_비케이(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """LG_비케이 계산 로직"""
        print("LG_비케이 계산 시작")
        
        for sheet_name in wb.sheetnames:
            if "_분석" in sheet_name:
                continue
                
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and self.is_pure_number(cell.value):
                        original_value = self.extract_number(cell.value)
                        if original_value != 0:
                            # 빨간색 텍스트는 음수로 변환
                            if self.is_red_text(cell):
                                original_value = -abs(original_value)
                                self.calculations_log.append(f"{sheet_name} {cell.coordinate}: 빨간색 -> 음수 변환: {original_value}")
                            
                            # 모든 상품에 +39
                            new_value = original_value + 39
                            self.calculations_log.append(f"{sheet_name} {cell.coordinate}: {original_value} + 39 = {new_value}")
                            
                            # 30 초과 시 +1
                            if new_value > 30:
                                new_value += 1
                                self.calculations_log.append(f"{sheet_name} {cell.coordinate}: 30 초과 +1 = {new_value}")
                            
                            # 60 초과 시 추가 +2
                            if new_value > 60:
                                new_value += 2
                                self.calculations_log.append(f"{sheet_name} {cell.coordinate}: 60 초과 +2 = {new_value}")
                            
                            # 빨간색 텍스트 여부 저장
                            was_red = self.is_red_text(cell)
                            
                            # 최종 계산
                            final_value = new_value * 10000
                            cell.value = final_value
                            
                            # 빨간색 텍스트였던 경우 색상 유지
                            if was_red:
                                from openpyxl.styles import Font
                                cell.font = Font(size=10, color="FF0000", bold=True)
        
        return wb
    
    def calculate_lg_엘에스(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """LG_엘에스 계산 로직"""
        print("LG_엘에스 계산 시작")
        
        for sheet_name in wb.sheetnames:
            if "_분석" in sheet_name:
                continue
                
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and self.is_pure_number(cell.value):
                        original_value = self.extract_number(cell.value)
                        if original_value != 0:
                            # 모든 상품에 2를 곱함
                            new_value = original_value * 2
                            self.calculations_log.append(f"{sheet_name} {cell.coordinate}: {original_value} * 2 = {new_value}")
                            
                            # 최종 계산
                            final_value = new_value * 10000
                            cell.value = final_value
        
        return wb
    
    def calculate_sk_나텔(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """SK_나텔 계산 로직"""
        print("SK_나텔 계산 시작")
        
        for sheet_name in wb.sheetnames:
            if "_분석" in sheet_name:
                continue
                
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and self.is_pure_number(cell.value):
                        original_value = self.extract_number(cell.value)
                        if original_value != 0:
                            # 최종 계산 (1000 곱하기)
                            final_value = original_value * 1000
                            cell.value = final_value
                            self.calculations_log.append(f"{sheet_name} {cell.coordinate}: {original_value} * 1000 = {final_value}")
        
        return wb
    
    def calculate_sk_상상(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """SK_상상 계산 로직"""
        print("SK_상상 계산 시작")
        
        for sheet_name in wb.sheetnames:
            if "_분석" in sheet_name:
                continue
                
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and self.is_pure_number(cell.value):
                        original_value = self.extract_number(cell.value)
                        if original_value != 0:
                            # 최종 계산 (1000 곱하기)
                            final_value = original_value * 1000
                            cell.value = final_value
                            self.calculations_log.append(f"{sheet_name} {cell.coordinate}: {original_value} * 1000 = {final_value}")
        
        return wb
    
    def calculate_sk_윤텔(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """SK_윤텔 계산 로직"""
        print("SK_윤텔 계산 시작")
        
        for sheet_name in wb.sheetnames:
            if "_분석" in sheet_name:
                continue
                
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and self.is_pure_number(cell.value):
                        original_value = self.extract_number(cell.value)
                        if original_value != 0:
                            # 최종 계산 (10000 곱하기)
                            final_value = original_value * 10000
                            cell.value = final_value
                            self.calculations_log.append(f"{sheet_name} {cell.coordinate}: {original_value} * 10000 = {final_value}")
        
        return wb
    
    def calculate_sk_텔컴(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """SK_텔컴 계산 로직"""
        print("SK_텔컴 계산 시작")
        
        for sheet_name in wb.sheetnames:
            if "_분석" in sheet_name:
                continue
                
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and self.is_pure_number(cell.value):
                        original_value = self.extract_number(cell.value)
                        if original_value != 0:
                            # 최종 계산 (1000 곱하기)
                            final_value = original_value * 1000
                            cell.value = final_value
                            self.calculations_log.append(f"{sheet_name} {cell.coordinate}: {original_value} * 1000 = {final_value}")
        
        return wb
    
    def calculate_sk_케이(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """SK_케이 계산 로직"""
        print("SK_케이 계산 시작")
        
        for sheet_name in wb.sheetnames:
            if "_분석" in sheet_name:
                continue
                
            ws = wb[sheet_name]
            
            # 모든 셀에 대해 단순 계산
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and self.is_pure_number(cell.value):
                        original_value = self.extract_number(cell.value)
                        if original_value != 0:
                            # 최종 계산 (10000 곱하기)
                            final_value = original_value * 10000
                            cell.value = final_value
                            self.calculations_log.append(f"{sheet_name} {cell.coordinate}: {original_value} * 10000 = {final_value}")
        
        return wb
    
    def calculate_sk_대교(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """SK_대교 계산 로직 - 단순 곱셈"""
        print("SK_대교 계산 시작")

        for sheet_name in wb.sheetnames:
            if "_분석" in sheet_name:
                continue

            ws = wb[sheet_name]

            # 모든 셀을 순회하며 계산
            calculated_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    if self.is_pure_number(cell.value):
                        original_value = self.extract_number(cell.value)
                        if original_value != 0:
                            # 최종 계산 (10000 곱하기)
                            final_value = original_value * 10000
                            cell.value = final_value
                            calculated_count += 1
                            self.calculations_log.append(f"{sheet_name} {cell.coordinate}: {original_value} * 10000 = {final_value}")

        print(f"SK_대교 계산 완료: {calculated_count}개 셀 계산됨")
        return wb

    def calculate_번개폰(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        """번개폰 계산 로직 - 추가 계산 없이 원본 값 유지"""
        print("번개폰 계산 시작")

        for sheet_name in wb.sheetnames:
            if "_분석" in sheet_name:
                continue

            ws = wb[sheet_name]

            # 모든 셀을 순회하며 원본 값 유지 (추가 계산 없음)
            calculated_count = 0
            for row in ws.iter_rows():
                for cell in row:
                    if self.is_pure_number(cell.value):
                        original_value = self.extract_number(cell.value)
                        if original_value != 0:
                            # 번개폰은 추가 계산 없이 원본 값 그대로 유지
                            cell.value = original_value
                            calculated_count += 1
                            self.calculations_log.append(f"{sheet_name} {cell.coordinate}: {original_value} (변경 없음)")

        print(f"번개폰 계산 완료: {calculated_count}개 셀 처리됨 (값 변경 없음)")
        return wb
    
    
    def process_excel_file(self, file_path: str) -> str:
        """Excel 파일 처리 및 계산 수행"""
        print(f"\n파일 처리 시작: {file_path}")
        
        # 파일명에서 대리점명 추출 (Unicode 정규화)
        file_name = os.path.basename(file_path)
        file_name = unicodedata.normalize('NFC', file_name)  # Unicode 정규화
        
        # 파일명에서 확장자와 접미사 제거
        dealer_name = file_name.replace('.xlsx', '').replace('_tables', '').replace('table_', '')
        
        # 날짜 패턴 제거 (예: 250708_)
        dealer_name = re.sub(r'^\d{6}_', '', dealer_name)
        
        # Excel 파일 로드
        file_path = Path(file_path)
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        # 계산 로그 초기화
        self.calculations_log = []
        
        # 대리점별 계산 수행
        if dealer_name == "KT_더블유":
            wb = self.calculate_kt_더블유(wb)
        elif dealer_name == "KT_맥스":
            wb = self.calculate_kt_맥스(wb)
        elif dealer_name == "LG_비케이":
            wb = self.calculate_lg_비케이(wb)
        elif dealer_name == "LG_엘에스":
            wb = self.calculate_lg_엘에스(wb)
        elif dealer_name == "SK_나텔":
            wb = self.calculate_sk_나텔(wb)
        elif dealer_name == "SK_상상":
            wb = self.calculate_sk_상상(wb)
        elif dealer_name == "SK_윤텔":
            wb = self.calculate_sk_윤텔(wb)
        elif dealer_name == "SK_케이":
            wb = self.calculate_sk_케이(wb)
        elif dealer_name == "SK_텔컴":
            wb = self.calculate_sk_텔컴(wb)
        elif dealer_name == "SK_대교":
            wb = self.calculate_sk_대교(wb)
        elif dealer_name == "LG_비케이2":
            # LG_비케이2는 LG_비케이와 동일한 로직 사용
            print("LG_비케이2는 LG_비케이와 동일한 계산 로직 적용")
            wb = self.calculate_lg_비케이(wb)
        elif "번개폰" in dealer_name:
            # 번개폰 파일은 추가 계산 없음
            print(f"번개폰 파일 감지: {dealer_name}")
            wb = self.calculate_번개폰(wb)
        elif "애플사전예약" in dealer_name or "apple" in dealer_name.lower():
            # 애플 사전예약 파일은 별도 처리
            print(f"애플 사전예약 파일 감지: {dealer_name}")
            from src.ocr.apple_preorder_calculator import calculate_apple_preorder
            output_file = calculate_apple_preorder(str(file_path))
            return output_file
        else:
            print(f"알 수 없는 대리점: {dealer_name}")
            print("기본 SK 계산 로직 적용")
            wb = self.calculate_sk_케이(wb)
        
        # 계산된 파일 저장
        output_file = file_path.with_name(file_path.stem + '_calculated.xlsx')
        wb.save(output_file)
        
        # 계산 로그 저장
        log_file = file_path.with_name(file_path.stem + '_log.txt')
        with open(log_file, 'w', encoding='utf-8') as f:
            f.write(f"계산 로그 - {dealer_name}\n")
            f.write(f"총 {len(self.calculations_log)}개 셀 계산됨\n\n")
            for log in self.calculations_log:
                f.write(log + '\n')
        
        print(f"계산 완료: {output_file}")
        print(f"로그 파일: {log_file}")
        print(f"총 {len(self.calculations_log)}개 셀 계산됨")
        
        return output_file


def main():
    """메인 함수"""
    import sys
    
    if len(sys.argv) < 2:
        print("사용법: python unit_price_calculator.py <엑셀파일경로>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(f"파일을 찾을 수 없습니다: {file_path}")
        sys.exit(1)
    
    calculator = UnitPriceCalculator()
    calculator.process_excel_file(file_path)


if __name__ == "__main__":
    main()