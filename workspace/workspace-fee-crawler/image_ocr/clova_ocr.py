#!/usr/bin/env python3
from pathlib import Path
from typing import List, Optional, Dict, Union
import typer
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TimeRemainingColumn, MofNCompleteColumn
from rich.prompt import Prompt, Confirm
from rich.syntax import Syntax
import json
from datetime import datetime
import pytz
import requests
import uuid
import time
import base64
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl.cell
import os
import cv2
import numpy as np
from collections import Counter


# 터미널 UI 초기화
app = typer.Typer(help="네이버 클로바 OCR 터미널 UI")
console = Console()

# API 설정
CONFIG_DIR = "/Users/jacob_athometrip/Desktop/dev/nofee/workspace_nofee/config"

def load_ocr_api_config():
    """CLOVA OCR API 설정 로드"""
    # clova_ocr_api_key.json 파일 찾기 (old가 붙지 않은 파일 우선)
    config_files = [
        os.path.join(CONFIG_DIR, "clova_ocr_api_key.json"),
        os.path.join(CONFIG_DIR, "clova_ocr_api_key_old.json"),
    ]

    for config_file in config_files:
        if os.path.exists(config_file):
            with open(config_file, 'r') as f:
                config = json.load(f)
                console.print(f"✅ CLOVA OCR API 설정 로드: {config_file}")
                return config['api_url'], config['secret_key']

    raise FileNotFoundError(f"CLOVA OCR API 키 파일을 찾을 수 없습니다: {CONFIG_DIR}")

API_URL, SECRET_KEY = load_ocr_api_config()


class ClovaOCRClient:
    """네이버 클로바 OCR API 클라이언트"""
    
    def __init__(self, api_url: str, secret_key: str):
        """
        ClovaOCR 클라이언트 초기화
        
        Args:
            api_url: CLOVA OCR API Gateway Invoke URL
            secret_key: API 인증용 Secret Key
        """
        self.api_url = api_url
        self.secret_key = secret_key
        
    def _create_request_body(self, images: List[Union[str, Path]], 
                           request_id: Optional[str] = None,
                           lang: str = "ko",
                           enable_table_detection: bool = False) -> Dict:
        """
        OCR 요청 본문 생성
        
        Args:
            images: 이미지 파일 경로 리스트
            request_id: 요청 ID (선택사항)
            lang: 언어 코드 (기본값: ko)
            enable_table_detection: 표 감지 활성화 여부
            
        Returns:
            요청 본문 딕셔너리
        """
        if not request_id:
            request_id = str(uuid.uuid4())
            
        image_list = []
        for idx, image_path in enumerate(images):
            if isinstance(image_path, str):
                image_path = Path(image_path)
                
            if not image_path.exists():
                raise FileNotFoundError(f"이미지 파일을 찾을 수 없습니다: {image_path}")
                
            # 이미지를 base64로 인코딩
            with open(image_path, 'rb') as f:
                image_data = base64.b64encode(f.read()).decode('utf-8')
                
            image_list.append({
                "format": image_path.suffix[1:].lower(),  # 확장자에서 . 제거
                "data": image_data,
                "name": f"image_{idx}"
            })
            
        request_body = {
            "version": "V2",
            "requestId": request_id,
            "timestamp": int(time.time() * 1000),
            "lang": lang,
            "images": image_list
        }

        # 표 감지는 항상 활성화 (General API는 기본적으로 표 감지를 지원)
        if enable_table_detection:
            request_body["enableTableDetection"] = True
        
        return request_body
        
    def recognize(self, image_paths: Union[str, Path, List[Union[str, Path]]], 
                  **kwargs) -> Dict:
        """
        OCR 인식 수행
        
        Args:
            image_paths: 이미지 파일 경로 또는 경로 리스트
            **kwargs: 추가 옵션 (lang, enable_table_detection 등)
            
        Returns:
            OCR 인식 결과
        """
        # 단일 이미지 경로를 리스트로 변환
        if isinstance(image_paths, (str, Path)):
            image_paths = [image_paths]
            
        # 요청 본문 생성
        request_body = self._create_request_body(image_paths, **kwargs)
        
        # API 요청 헤더
        headers = {
            'X-OCR-SECRET': self.secret_key,
            'Content-Type': 'application/json'
        }
        
        # API 호출
        try:
            response = requests.post(
                self.api_url,
                headers=headers,
                data=json.dumps(request_body),
                timeout=30,
                verify=True  # SSL 인증서 검증
            )
            response.raise_for_status()
            
            return response.json()
            
        except requests.exceptions.RequestException as e:
            raise Exception(f"OCR API 요청 실패: {str(e)}")
            
    def extract_text(self, ocr_result: Dict) -> List[str]:
        """
        OCR 결과에서 텍스트만 추출
        
        Args:
            ocr_result: OCR API 응답 결과
            
        Returns:
            추출된 텍스트 리스트
        """
        texts = []
        
        if 'images' in ocr_result:
            for image in ocr_result['images']:
                if 'fields' in image:
                    for field in image['fields']:
                        if 'inferText' in field:
                            texts.append(field['inferText'])
                            
        return texts
        
    def extract_tables(self, ocr_result: Dict) -> List[List[List[str]]]:
        """
        OCR 결과에서 표 데이터 추출
        
        Args:
            ocr_result: OCR API 응답 결과
            
        Returns:
            추출된 표 데이터 (표 리스트 > 행 리스트 > 셀 리스트)
        """
        tables = []
        
        if 'images' in ocr_result:
            for image in ocr_result['images']:
                if 'tables' in image:
                    for table in image['tables']:
                        table_data = []
                        if 'cells' in table:
                            # 셀을 행과 열로 정리
                            max_row = max(cell['rowIndex'] for cell in table['cells'])
                            max_col = max(cell['columnIndex'] for cell in table['cells'])
                            
                            # 빈 테이블 생성
                            table_matrix = [['' for _ in range(max_col + 1)] 
                                          for _ in range(max_row + 1)]
                            
                            # 셀 데이터 채우기
                            for cell in table['cells']:
                                row_idx = cell['rowIndex']
                                col_idx = cell['columnIndex']
                                cell_text = cell.get('cellTextLines', [])
                                table_matrix[row_idx][col_idx] = ' '.join(
                                    line.get('inferText', '') for line in cell_text
                                )
                                
                            tables.append(table_matrix)
                            
        return tables
    
    def extract_tables_with_coordinates(self, ocr_result: Dict) -> List[Dict]:
        """
        OCR 결과에서 표 데이터를 좌표 정보와 함께 추출
        
        Args:
            ocr_result: OCR API 응답 결과
            
        Returns:
            좌표 정보가 포함된 표 데이터 리스트
        """
        tables_with_coords = []
        
        if 'images' in ocr_result:
            for image in ocr_result['images']:
                if 'tables' in image:
                    for table in image['tables']:
                        table_info = {
                            'cells': [],
                            'max_row': 0,
                            'max_col': 0
                        }
                        
                        if 'cells' in table:
                            table_info['max_row'] = max(cell['rowIndex'] + cell.get('rowSpan', 1) - 1 
                                                       for cell in table['cells'])
                            table_info['max_col'] = max(cell['columnIndex'] + cell.get('columnSpan', 1) - 1 
                                                       for cell in table['cells'])
                            
                            for cell in table['cells']:
                                cell_data = {
                                    'rowIndex': cell['rowIndex'],
                                    'columnIndex': cell['columnIndex'],
                                    'rowSpan': cell.get('rowSpan', 1),
                                    'columnSpan': cell.get('columnSpan', 1),
                                    'text': '',
                                    'confidence': cell.get('inferConfidence', 0),
                                    'boundingPoly': cell.get('boundingPoly', {})
                                }
                                
                                # 셀 텍스트 추출
                                cell_texts = []
                                for line in cell.get('cellTextLines', []):
                                    line_words = []
                                    for word in line.get('cellWords', []):
                                        line_words.append(word.get('inferText', ''))
                                    if line_words:
                                        cell_texts.append(' '.join(line_words))
                                    else:
                                        # cellWords가 없는 경우 inferText 직접 사용
                                        cell_texts.append(line.get('inferText', ''))
                                
                                cell_data['text'] = ' '.join(cell_texts)
                                table_info['cells'].append(cell_data)
                            
                        tables_with_coords.append(table_info)
                            
        return tables_with_coords
    
    def extract_cell_colors(self, image_path: Union[str, Path], ocr_result: Dict) -> Dict[str, str]:
        """
        이미지에서 각 셀의 배경색 추출
        
        Args:
            image_path: 이미지 파일 경로
            ocr_result: OCR 결과
            
        Returns:
            셀 인덱스와 색상 코드 매핑 딕셔너리
        """
        if isinstance(image_path, str):
            image_path = Path(image_path)
            
        # 이미지 읽기
        img = cv2.imread(str(image_path))
        if img is None:
            return {}
            
        # BGR을 RGB로 변환
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        
        cell_colors = {}
        
        if 'images' in ocr_result:
            for image in ocr_result['images']:
                if 'tables' in image:
                    for table in image['tables']:
                        if 'cells' in table:
                            for cell in table['cells']:
                                if 'boundingPoly' in cell and 'vertices' in cell['boundingPoly']:
                                    vertices = cell['boundingPoly']['vertices']
                                    if len(vertices) >= 4:
                                        # 셀의 중심점 계산
                                        center_x = int(sum(v['x'] for v in vertices) / len(vertices))
                                        center_y = int(sum(v['y'] for v in vertices) / len(vertices))
                                        
                                        # 셀 영역 내 여러 지점 샘플링
                                        x1 = int(min(v['x'] for v in vertices))
                                        y1 = int(min(v['y'] for v in vertices))
                                        x2 = int(max(v['x'] for v in vertices))
                                        y2 = int(max(v['y'] for v in vertices))
                                        
                                        # 셀 영역 내에서 색상 샘플링
                                        sample_points = []
                                        margin = 5  # 가장자리에서 5픽셀 안쪽
                                        
                                        for y in range(y1 + margin, min(y2 - margin, img_rgb.shape[0]), 10):
                                            for x in range(x1 + margin, min(x2 - margin, img_rgb.shape[1]), 10):
                                                if 0 <= y < img_rgb.shape[0] and 0 <= x < img_rgb.shape[1]:
                                                    sample_points.append(img_rgb[y, x])
                                        
                                        if sample_points:
                                            # 가장 많이 나타나는 색상 찾기
                                            colors = [tuple(color) for color in sample_points]
                                            color_counter = Counter(colors)
                                            dominant_color = color_counter.most_common(1)[0][0]
                                            
                                            # RGB를 HEX로 변환
                                            hex_color = '{:02X}{:02X}{:02X}'.format(
                                                dominant_color[0], dominant_color[1], dominant_color[2]
                                            )
                                            
                                            row_idx = cell.get('rowIndex', 0)
                                            col_idx = cell.get('columnIndex', 0)
                                            cell_key = f"{row_idx}_{col_idx}"
                                            cell_colors[cell_key] = hex_color
        
        return cell_colors
    
    def extract_text_colors(self, image_path: Union[str, Path], ocr_result: Dict) -> Dict[str, str]:
        """
        이미지에서 각 셀의 텍스트 색상 추출
        
        Args:
            image_path: 이미지 파일 경로
            ocr_result: OCR 결과
            
        Returns:
            셀 인덱스와 텍스트 색상 코드 매핑 딕셔너리
        """
        if isinstance(image_path, str):
            image_path = Path(image_path)
            
        # 이미지 읽기
        img = cv2.imread(str(image_path))
        if img is None:
            return {}
            
        text_colors = {}
        
        if 'images' in ocr_result:
            for image in ocr_result['images']:
                if 'tables' in image:
                    for table in image['tables']:
                        if 'cells' in table:
                            for cell in table['cells']:
                                row_idx = cell.get('rowIndex', 0)
                                col_idx = cell.get('columnIndex', 0)
                                
                                # 셀 내의 텍스트 라인들 확인
                                if 'cellTextLines' in cell and cell['cellTextLines']:
                                    cell_has_red = False
                                    
                                    for text_line in cell['cellTextLines']:
                                        if 'boundingPoly' in text_line and 'vertices' in text_line['boundingPoly']:
                                            vertices = text_line['boundingPoly']['vertices']
                                            if len(vertices) >= 4:
                                                x1 = int(min(v['x'] for v in vertices))
                                                y1 = int(min(v['y'] for v in vertices))
                                                x2 = int(max(v['x'] for v in vertices))
                                                y2 = int(max(v['y'] for v in vertices))
                                                
                                                # 영역이 이미지 범위 내에 있는지 확인
                                                x1 = max(0, x1)
                                                y1 = max(0, y1)
                                                x2 = min(img.shape[1], x2)
                                                y2 = min(img.shape[0], y2)
                                                
                                                if x2 > x1 and y2 > y1:
                                                    # 텍스트 영역의 색상 감지
                                                    region = img[y1:y2, x1:x2]
                                                    if region.size > 0:
                                                        # 그레이스케일로 변환하여 텍스트 위치 찾기
                                                        gray = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
                                                        
                                                        # Otsu's thresholding으로 텍스트 영역 찾기
                                                        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
                                                        
                                                        # 텍스트 픽셀 위치 찾기
                                                        text_pixels_coords = np.where(binary == 255)
                                                        
                                                        if len(text_pixels_coords[0]) > 0:
                                                            # 텍스트 픽셀의 색상 수집
                                                            red_count = 0
                                                            total_count = 0
                                                            
                                                            for i in range(len(text_pixels_coords[0])):
                                                                y = text_pixels_coords[0][i]
                                                                x = text_pixels_coords[1][i]
                                                                # BGR to RGB
                                                                b, g, r = region[y, x]
                                                                r, g, b = int(r), int(g), int(b)
                                                                
                                                                total_count += 1
                                                                
                                                                # 빨간색 판별 (실제 측정값 기반)
                                                                if r >= 250 and g < 170 and b < 220:
                                                                    red_count += 1
                                                                elif r >= 200 and g < 100 and b < 100 and r > g * 2 and r > b * 2:
                                                                    red_count += 1
                                                            
                                                            # 30% 이상이 빨간색이면 빨간색으로 판단
                                                            if total_count > 0 and red_count > total_count * 0.3:
                                                                cell_has_red = True
                                    
                                    # 셀에 빨간색 텍스트가 하나라도 있으면 빨간색으로 표시
                                    cell_key = f"{row_idx}_{col_idx}"
                                    if cell_has_red:
                                        text_colors[cell_key] = "red"
                                    else:
                                        text_colors[cell_key] = "black"
        
        return text_colors


# 결과 포맷터 클래스
class OCRResultFormatter:
    """OCR 결과를 다양한 형식으로 포맷팅"""
    
    @staticmethod
    def to_json(ocr_result: Dict, pretty: bool = True) -> str:
        """JSON 형식으로 변환"""
        if pretty:
            return json.dumps(ocr_result, ensure_ascii=False, indent=2)
        return json.dumps(ocr_result, ensure_ascii=False)
        
    @staticmethod
    def to_text(ocr_result: Dict) -> str:
        """텍스트 형식으로 변환"""
        client = ClovaOCRClient("", "")  # URL과 키는 불필요
        texts = client.extract_text(ocr_result)
        return '\n'.join(texts)
        


class OCRTerminalUI:
    """OCR 터미널 UI 클래스"""
    
    def __init__(self):
        self.client = ClovaOCRClient(API_URL, SECRET_KEY)

        # PathManager를 사용한 중앙화된 경로 관리
        from shared_config.config.paths import PathManager
        pm = PathManager()

        # 입력 이미지 폴더
        self.ocr_folder = pm.ocr_input_dir

        # 출력 폴더 (Latest + Archive 패턴)
        kst = pytz.timezone('Asia/Seoul')
        now_kst = datetime.now(kst)
        date_str = now_kst.strftime("%Y%m%d")

        # Latest 폴더 (항상 최신 결과)
        self.output_folder_latest = pm.ocr_latest_dir

        # Archive 폴더 (날짜별 히스토리)
        self.output_folder_archive = pm.ocr_archive_dir / date_str
        self.output_folder_archive.mkdir(parents=True, exist_ok=True)

        # 기본 출력 폴더는 archive (히스토리 보존)
        self.output_folder = self.output_folder_archive

        console.print(f"[green]📁 입력 폴더: {self.ocr_folder}[/green]")
        console.print(f"[green]📁 출력 폴더 (Archive): {self.output_folder}[/green]")
        console.print(f"[green]📁 출력 폴더 (Latest): {self.output_folder_latest}[/green]")
        
    def show_banner(self):
        """배너 표시"""
        banner = """
╔═══════════════════════════════════════════════════════════╗
║                   네이버 클로바 OCR 도구                    ║
║                  이미지에서 텍스트 추출하기                 ║
╚═══════════════════════════════════════════════════════════╝
        """
        console.print(Panel(banner, style="bold blue"))
        
    def list_images(self) -> List[Path]:
        """이미지 파일 목록 조회"""
        if not self.ocr_folder.exists():
            console.print(f"[red]폴더를 찾을 수 없습니다: {self.ocr_folder}[/red]")
            return []
            
        image_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.bmp'}
        images = [f for f in self.ocr_folder.iterdir() 
                 if f.is_file() and f.suffix.lower() in image_extensions]
        
        return sorted(images)
        
    def show_image_list(self, images: List[Path]) -> Optional[List[Path]]:
        """이미지 목록을 표시하고 선택받기"""
        if not images:
            console.print("[yellow]이미지 파일이 없습니다.[/yellow]")
            return None
            
        table = Table(title="이미지 파일 목록", show_header=True, header_style="bold magenta")
        table.add_column("번호", style="cyan", width=6)
        table.add_column("파일명", style="green")
        table.add_column("크기", style="yellow", justify="right")
        
        for idx, img in enumerate(images, 1):
            size = img.stat().st_size / 1024  # KB
            table.add_row(str(idx), img.name, f"{size:.1f} KB")
            
        console.print(table)
        
        # 선택 받기
        console.print("\n[bold]옵션:[/bold]")
        console.print("  - 번호 입력: 단일 파일 선택 (예: 1)")
        console.print("  - 범위 입력: 여러 파일 선택 (예: 1-3)")
        console.print("  - 쉼표 구분: 개별 선택 (예: 1,3,5)")
        console.print("  - 'all': 모든 파일 선택")
        console.print("  - 'q': 종료")
        
        choice = Prompt.ask("\n선택", default="all")
        
        if choice.lower() == 'q':
            return None
        elif choice.lower() == 'all':
            return images
        else:
            selected = []
            try:
                if '-' in choice:  # 범위 선택
                    start, end = map(int, choice.split('-'))
                    selected = images[start-1:end]
                elif ',' in choice:  # 개별 선택
                    indices = [int(x.strip()) - 1 for x in choice.split(',')]
                    selected = [images[i] for i in indices if 0 <= i < len(images)]
                else:  # 단일 선택
                    idx = int(choice) - 1
                    if 0 <= idx < len(images):
                        selected = [images[idx]]
                        
                return selected if selected else None
                
            except (ValueError, IndexError):
                console.print("[red]잘못된 입력입니다.[/red]")
                return None
                
    def process_image(self, image_path: Path, enable_table: bool = False, show_progress: bool = True) -> Optional[dict]:
        """단일 이미지 OCR 처리"""
        try:
            if show_progress:
                with Progress(
                    SpinnerColumn(),
                    TextColumn("[progress.description]{task.description}"),
                    transient=True,
                ) as progress:
                    task = progress.add_task(f"[cyan]{image_path.name} 처리 중...", total=None)
                    result = self.client.recognize(image_path, lang="ko", enable_table_detection=enable_table)
                    progress.remove_task(task)
            else:
                result = self.client.recognize(image_path, lang="ko", enable_table_detection=enable_table)
                
            return result
            
        except Exception as e:
            console.print(f"[red]오류 발생: {str(e)}[/red]")
            return None
            
    def display_result(self, image_name: str, result: dict, show_json: bool = False):
        """OCR 결과 표시"""
        console.print(f"\n[bold blue]═══ {image_name} 결과 ═══[/bold blue]")
        
        if show_json:
            # JSON 형식으로 표시
            json_str = OCRResultFormatter.to_json(result)
            syntax = Syntax(json_str, "json", theme="monokai", line_numbers=True)
            console.print(Panel(syntax, title="JSON 결과"))
        else:
            # 텍스트만 표시
            texts = self.client.extract_text(result)
            if texts:
                text_panel = Panel(
                    "\n".join(texts),
                    title="추출된 텍스트",
                    border_style="green"
                )
                console.print(text_panel)
            else:
                console.print("[yellow]텍스트를 찾을 수 없습니다.[/yellow]")
                
            # 표 데이터가 있으면 표시
            tables = self.client.extract_tables(result)
            if tables:
                for idx, table_data in enumerate(tables, 1):
                    table = Table(title=f"표 {idx}", show_header=False, show_lines=True)
                    
                    # 컬럼 추가
                    max_cols = max(len(row) for row in table_data) if table_data else 0
                    for _ in range(max_cols):
                        table.add_column()
                        
                    # 행 추가
                    for row in table_data:
                        table.add_row(*row)
                        
                    console.print(table)
                    
    def save_results(self, results: dict, save_individually: bool = True, image_paths: dict = None):
        """결과 저장 (JSON 및 Excel 형식)"""
        saved_files = []
        
        if save_individually:
            # 각 이미지별로 개별 파일 저장
            for image_name, result in results.items():
                base_name = Path(image_name).stem  # 확장자 제거
                
                # 개별 JSON 파일 저장
                json_filename = self.output_folder / f"{base_name}.json"
                individual_result = {image_name: result}
                with open(json_filename, 'w', encoding='utf-8') as f:
                    json.dump(individual_result, f, ensure_ascii=False, indent=2)
                console.print(f"[green]JSON 파일 저장: {json_filename}[/green]")
                saved_files.append(json_filename)
                
                # 개별 Excel 파일 생성
                excel_filename = self.output_folder / f"{base_name}.xlsx"
                tables_filename = self.output_folder / f"{base_name}_tables.xlsx"
                
                # 이미지 경로 전달
                individual_image_paths = None
                if image_paths and image_name in image_paths:
                    individual_image_paths = {image_name: image_paths[image_name]}
                
                # Excel 파일 저장 (개별)
                self._save_individual_excel(individual_result, excel_filename, tables_filename, individual_image_paths)
                saved_files.append(excel_filename)
                if tables_filename.exists():
                    saved_files.append(tables_filename)
                    # 색상 추출 대상 파일인지 확인
                    if any(keyword in image_name for keyword in ["맥스", "더블유", "비케이", "엘에스", "나텔", "상상", "윤텔", "케이"]):
                        console.print(f"[magenta]🎨 {image_name}: 색상 정보가 포함된 표 데이터 저장 완료[/magenta]")
            
            # 전체 결과도 하나의 파일로 저장
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            all_json_filename = self.output_folder / f"all_results_{timestamp}.json"
            with open(all_json_filename, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            console.print(f"[green]전체 JSON 파일 저장: {all_json_filename}[/green]")
            saved_files.append(all_json_filename)
            
        else:
            # 기존 방식 (단일 파일로 저장)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # JSON 파일 저장
            json_filename = self.output_folder / f"ocr_results_{timestamp}.json"
            with open(json_filename, 'w', encoding='utf-8') as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            console.print(f"[green]JSON 파일 저장: {json_filename}[/green]")
            saved_files.append(json_filename)
            
            # Excel 파일 저장
            excel_filename = self.output_folder / f"ocr_results_{timestamp}.xlsx"
            excel_tables_filename = self.output_folder / f"ocr_tables_{timestamp}.xlsx"
            
            self._save_combined_excel(results, excel_filename, excel_tables_filename)
            saved_files.append(excel_filename)
            if excel_tables_filename.exists():
                saved_files.append(excel_tables_filename)
        
        return saved_files
    
    def _save_individual_excel(self, result: dict, excel_filename: Path, tables_filename: Path, image_paths: dict = None):
        """개별 이미지의 Excel 파일 저장"""
        self._save_combined_excel(result, excel_filename, tables_filename, image_paths)
    
    def _save_combined_excel(self, results: dict, excel_filename: Path, excel_tables_filename: Path, image_paths: dict = None):
        """Excel 파일 저장 로직"""
        # 메인 Excel 파일 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "OCR 결과"
        
        # 스타일 정의
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더 작성
        headers = ["파일명", "텍스트", "신뢰도", "좌표(x1,y1,x2,y2)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 데이터 작성
        row_num = 2
        for image_name, result in results.items():
            if 'images' in result:
                for image in result['images']:
                    # 텍스트 필드 처리
                    if 'fields' in image:
                        for field in image['fields']:
                            text = field.get('inferText', '')
                            confidence = field.get('inferConfidence', '')
                            
                            # 좌표 정보 추출
                            coords = ""
                            if 'boundingPoly' in field and 'vertices' in field['boundingPoly']:
                                vertices = field['boundingPoly']['vertices']
                                if len(vertices) >= 4:
                                    coords = f"{vertices[0]['x']},{vertices[0]['y']},{vertices[2]['x']},{vertices[2]['y']}"
                            
                            # 데이터 쓰기
                            ws.cell(row=row_num, column=1, value=image_name).border = border
                            ws.cell(row=row_num, column=2, value=text).border = border
                            ws.cell(row=row_num, column=3, value=confidence).border = border
                            ws.cell(row=row_num, column=4, value=coords).border = border
                            row_num += 1
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_filename)
        console.print(f"[green]Excel 파일 저장: {excel_filename}[/green]")
        
        # 표 데이터를 별도 Excel 파일로 저장 (좌표 기반 정교한 변환)
        table_wb = Workbook()
        table_wb.remove(table_wb.active)  # 기본 시트 제거
        
        has_tables = False
        for image_name, result in results.items():
            tables_with_coords = self.client.extract_tables_with_coordinates(result)
            if tables_with_coords:
                has_tables = True
                
                # 색상 추출이 필요한 파일인지 확인 (모든 파일에 대해 색상 추출)
                should_extract_colors = True  # 모든 파일에 대해 색상 추출
                
                # 글자색 추출이 필요한 파일 확인 - 모든 파일에서 글자색 추출 시도
                text_color_required = True  # 모든 파일에서 글자색 추출
                cell_colors = {}
                text_colors = {}
                
                if should_extract_colors:
                    if image_paths and image_name in image_paths:
                        try:
                            console.print(f"[cyan]📊 {image_name}에서 셀 색상 추출 시작...[/cyan]")
                            cell_colors = self.client.extract_cell_colors(image_paths[image_name], result)
                            if cell_colors:
                                console.print(f"[green]✅ {image_name}: {len(cell_colors)}개 셀의 배경색 추출 완료[/green]")
                                # 색상 샘플 출력
                                sample_colors = list(cell_colors.items())[:5]
                                for cell_key, color in sample_colors:
                                    console.print(f"   - 셀 {cell_key}: #{color}")
                            else:
                                console.print(f"[yellow]⚠️  {image_name}: 배경색을 추출할 수 없습니다[/yellow]")
                            
                            # 글자색 추출 (LG_비케이의 경우)
                            if text_color_required:
                                console.print(f"[cyan]🖍️  {image_name}에서 글자색 추출 시작...[/cyan]")
                                try:
                                    text_colors = self.client.extract_text_colors(image_paths[image_name], result)
                                    if text_colors:
                                        console.print(f"[green]✅ {image_name}: {len(text_colors)}개 셀의 글자색 추출 완료[/green]")
                                        # 글자색 샘플 출력
                                        sample_text_colors = list(text_colors.items())[:5]
                                        for cell_key, color in sample_text_colors:
                                            console.print(f"   - 셀 {cell_key}: {color}")
                                    else:
                                        console.print(f"[yellow]⚠️  {image_name}: 글자색을 추출할 수 없습니다[/yellow]")
                                except Exception as e:
                                    console.print(f"[red]❌ {image_name}: 글자색 추출 중 오류 발생 - {str(e)}[/red]")
                                    
                        except Exception as e:
                            console.print(f"[red]❌ {image_name}: 색상 추출 중 오류 발생 - {str(e)}[/red]")
                    else:
                        console.print(f"[yellow]⚠️  {image_name}: 이미지 경로를 찾을 수 없어 색상 추출 건너뜁니다[/yellow]")
                
                # 각 이미지별로 시트 생성
                ws = table_wb.create_sheet(title=image_name[:31])  # Excel 시트명 31자 제한
                
                current_row = 1
                for table_idx, table_info in enumerate(tables_with_coords, 1):
                    # 표 제목
                    title_cell = ws.cell(row=current_row, column=1, value=f"표 {table_idx}")
                    title_cell.font = Font(bold=True, size=14, color="000080")
                    current_row += 2
                    
                    # 표 시작 위치
                    table_start_row = current_row
                    
                    # 각 셀을 올바른 위치에 배치
                    for cell_data in table_info['cells']:
                        row_idx = table_start_row + cell_data['rowIndex']
                        col_idx = cell_data['columnIndex'] + 1  # Excel은 1부터 시작
                        
                        # 셀 병합 처리
                        if cell_data['rowSpan'] > 1 or cell_data['columnSpan'] > 1:
                            end_row = row_idx + cell_data['rowSpan'] - 1
                            end_col = col_idx + cell_data['columnSpan'] - 1
                            ws.merge_cells(start_row=row_idx, start_column=col_idx,
                                         end_row=end_row, end_column=end_col)
                        
                        # 셀 데이터 입력
                        cell = ws.cell(row=row_idx, column=col_idx, value=cell_data['text'])
                        
                        # 셀 스타일 적용
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border
                        
                        # 색상 적용
                        cell_key = f"{cell_data['rowIndex']}_{cell_data['columnIndex']}"
                        
                        # 글자색 설정 (LG_비케이의 경우)
                        if text_colors and cell_key in text_colors:
                            text_color = text_colors[cell_key]
                            if text_color == "red":
                                cell.font = Font(size=10, color="FF0000")  # 빨간색
                            elif text_color == "black":
                                cell.font = Font(size=10, color="000000")  # 검정색
                            else:
                                cell.font = Font(size=10)  # 기본 폰트
                        else:
                            cell.font = Font(size=10)  # 기본 폰트
                        
                        # 배경색 적용
                        if cell_colors and cell_key in cell_colors:
                            hex_color = cell_colors[cell_key]
                            # 흰색이 아닌 경우에만 적용
                            if hex_color.upper() not in ['FFFFFF', 'FEFEFE', 'FDFDFD']:
                                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        # 신뢰도가 낮은 셀은 다른 색으로 표시 (색상이 없는 경우에만)
                        elif cell_data['confidence'] < 0.95 and not (cell_colors and cell_key in cell_colors):
                            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                    
                    # 표 전체에 테두리 적용
                    table_end_row = table_start_row + table_info['max_row']
                    table_end_col = table_info['max_col'] + 1
                    
                    for row in range(table_start_row, table_end_row + 1):
                        for col in range(1, table_end_col + 1):
                            cell = ws.cell(row=row, column=col)
                            # MergedCell은 값을 설정할 수 없으므로 건너뛰기
                            if not isinstance(cell, openpyxl.cell.MergedCell):
                                if cell.value is None:
                                    cell.value = ''
                                cell.border = border
                    
                    current_row = table_end_row + 3  # 표 사이 공백
                
                # 열 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # pandas DataFrame으로도 변환하여 분석 시트 추가
                df_sheet = table_wb.create_sheet(title=f"{image_name[:20]}_분석")
                
                # 모든 텍스트 필드를 DataFrame으로 변환
                text_data = []
                if 'images' in result:
                    for image in result['images']:
                        if 'fields' in image:
                            for field in image['fields']:
                                text = field.get('inferText', '')
                                confidence = field.get('inferConfidence', 0)
                                vertices = field.get('boundingPoly', {}).get('vertices', [])
                                
                                if vertices and len(vertices) >= 4:
                                    x1, y1 = vertices[0].get('x', 0), vertices[0].get('y', 0)
                                    x2, y2 = vertices[2].get('x', 0), vertices[2].get('y', 0)
                                    
                                    text_data.append({
                                        '텍스트': text,
                                        '신뢰도': confidence,
                                        'X1': x1,
                                        'Y1': y1,
                                        'X2': x2,
                                        'Y2': y2,
                                        '너비': x2 - x1,
                                        '높이': y2 - y1
                                    })
                
                if text_data:
                    df = pd.DataFrame(text_data)
                    # Y1 좌표 기준으로 정렬 (위에서 아래로)
                    df = df.sort_values(by=['Y1', 'X1'])
                    
                    # DataFrame을 Excel에 쓰기
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            cell = df_sheet.cell(row=r_idx, column=c_idx, value=value)
                            if r_idx == 1:  # 헤더
                                cell.font = header_font
                                cell.fill = header_fill
                            cell.border = border
                    
                    # 열 너비 자동 조정
                    for column in df_sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30)
                        df_sheet.column_dimensions[column_letter].width = adjusted_width
        
        if has_tables:
            table_wb.save(excel_tables_filename)
            console.print(f"[green]표 데이터 Excel 파일 저장: {excel_tables_filename}[/green]")
            
            # 색상 추출 결과 요약
            for image_name in results.keys():
                if any(keyword in image_name for keyword in ["맥스", "더블유", "비케이", "엘에스", "나텔", "상상", "윤텔", "케이"]):
                    console.print(f"[blue]📊 {image_name}: 색상 추출 대상 파일[/blue]")
            
            # 단가 계산 수행
            try:
                from .unit_price_calculator import UnitPriceCalculator
                console.print(f"\n[yellow]💰 단가 계산 시작...[/yellow]")
                calculator = UnitPriceCalculator()
                calculated_file = calculator.process_excel_file(excel_tables_filename)
                if calculated_file:
                    console.print(f"[green]✅ 단가 계산 완료: {calculated_file}[/green]")
                else:
                    console.print(f"[red]❌ 단가 계산 실패[/red]")
            except Exception as e:
                console.print(f"[red]❌ 단가 계산 중 오류 발생: {str(e)}[/red]")


@app.command()
def interactive():
    """대화형 모드로 OCR 실행"""
    ui = OCRTerminalUI()
    ui.show_banner()
    
    while True:
        # 이미지 목록 표시
        images = ui.list_images()
        selected_images = ui.show_image_list(images)
        
        if not selected_images:
            if Confirm.ask("\n종료하시겠습니까?", default=True):
                console.print("[yellow]프로그램을 종료합니다.[/yellow]")
                break
            continue
            
        # OCR 옵션 선택
        enable_table = Confirm.ask("\n표 감지를 활성화하시겠습니까?", default=False)
        show_json = Confirm.ask("JSON 형식으로 결과를 보시겠습니까?", default=False)
        
        # OCR 처리
        results = {}
        with console.status("[bold green]OCR 처리 중...") as status:
            for image in selected_images:
                status.update(f"[bold green]{image.name} 처리 중...")
                result = ui.process_image(image, enable_table)
                if result:
                    results[image.name] = result
                    ui.display_result(image.name, result, show_json)
                    
        # 결과 저장 여부
        if results and Confirm.ask("\n결과를 저장하시겠습니까?", default=True):
            ui.save_results(results)
            
        if not Confirm.ask("\n계속하시겠습니까?", default=True):
            break
            
    console.print("\n[bold blue]감사합니다! 👋[/bold blue]")


@app.command()
def quick(
    image: str = typer.Argument(..., help="이미지 파일 경로"),
    table: bool = typer.Option(False, "--table", "-t", help="표 감지 활성화"),
    save: bool = typer.Option(False, "--save", "-s", help="결과 저장")
):
    """빠른 OCR 실행 (단일 이미지)"""
    ui = OCRTerminalUI()
    
    # 경로 처리
    if not Path(image).is_absolute():
        image_path = ui.ocr_folder / image
    else:
        image_path = Path(image)
        
    if not image_path.exists():
        console.print(f"[red]파일을 찾을 수 없습니다: {image_path}[/red]")
        return
        
    # OCR 처리
    result = ui.process_image(image_path, table)
    if not result:
        return
        
    # 결과 표시 (텍스트만)
    texts = ui.client.extract_text(result)
    for text in texts:
        console.print(text)
        
    # 저장
    if save:
        ui.save_results({image_path.name: result}, image_paths={image_path.name: image_path})


@app.command()
def batch(
    pattern: str = typer.Argument("*.png", help="파일 패턴 (예: *.png, *KT*.png)"),
    table: bool = typer.Option(False, "--table", "-t", help="표 감지 활성화")
):
    """배치 OCR 실행 (여러 이미지)"""
    ui = OCRTerminalUI()
    
    # 패턴에 맞는 파일 찾기
    images = list(ui.ocr_folder.glob(pattern))
    
    if not images:
        console.print(f"[yellow]패턴에 맞는 파일이 없습니다: {pattern}[/yellow]")
        return
        
    console.print(f"[green]{len(images)}개 파일을 찾았습니다.[/green]")
    
    # OCR 처리
    results = {}
    image_paths = {}  # 이미지 경로 저장
    
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        MofNCompleteColumn(),
        TimeRemainingColumn()
    ) as progress:
        task = progress.add_task("[cyan]OCR 처리 중...", total=len(images))
        
        for image in images:
            # Progress 표시 중복 방지를 위해 show_progress=False
            result = ui.process_image(image, table, show_progress=False)
            if result:
                results[image.name] = result
                image_paths[image.name] = image  # 경로 저장
            progress.update(task, advance=1)
            
    # 결과 저장
    if results:
        saved_files = ui.save_results(results, image_paths=image_paths)
        console.print(f"\n[bold green]완료! {len(results)}개 파일 처리됨[/bold green]")
        console.print(f"[green]저장된 파일 수: {len(saved_files)}개[/green]")


@app.command()
def convert_json(
    json_file: str = typer.Argument(..., help="변환할 JSON 파일 경로"),
    output_dir: str = typer.Option(None, "--output", "-o", help="출력 디렉토리")
):
    """기존 JSON 파일을 Excel로 변환"""
    json_path = Path(json_file)
    if not json_path.exists():
        console.print(f"[red]파일을 찾을 수 없습니다: {json_path}[/red]")
        return
        
    # 출력 디렉토리 설정
    if output_dir:
        output_path = Path(output_dir)
    else:
        output_path = json_path.parent
    output_path.mkdir(exist_ok=True)
    
    # JSON 파일 읽기
    console.print(f"[cyan]JSON 파일 읽는 중: {json_path}[/cyan]")
    with open(json_path, 'r', encoding='utf-8') as f:
        results = json.load(f)
    
    # UI 인스턴스 생성 (결과 저장용)
    ui = OCRTerminalUI()
    ui.output_folder = output_path
    
    # 결과 저장 (이미지 경로 없이 변환)
    with console.status("[bold green]Excel 파일로 변환 중..."):
        saved_files = ui.save_results(results, save_individually=True, image_paths=None)
    
    console.print("\n[bold green]변환 완료![/bold green]")
    console.print(f"저장된 파일:")
    for file in saved_files:
        console.print(f"  - {file}")


if __name__ == "__main__":
    app()