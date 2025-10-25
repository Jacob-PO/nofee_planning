import pandas as pd
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import load_workbook
import glob

def hex_to_rgb(hex_color):
    """16진수 색상을 RGB 딕셔너리로 변환"""
    if not hex_color or hex_color == '00000000':
        return None
    
    # ARGB 형식인 경우 앞의 2자리(Alpha) 제거
    if len(hex_color) == 8:
        hex_color = hex_color[2:]
    
    try:
        r = int(hex_color[0:2], 16) / 255.0
        g = int(hex_color[2:4], 16) / 255.0
        b = int(hex_color[4:6], 16) / 255.0
        return {"red": r, "green": g, "blue": b}
    except:
        return None

def update_google_sheet_with_colors(excel_file_path, spreadsheet_id, sheet_name, service):
    """엑셀 파일의 데이터와 색상을 Google Sheets에 업데이트"""
    
    # 1. 엑셀 파일 읽기
    print(f"엑셀 파일 읽는 중: {excel_file_path}")
    df = pd.read_excel(excel_file_path)
    
    # 2. 엑셀 파일에서 색상 정보 추출
    wb = load_workbook(excel_file_path, data_only=True)
    ws = wb.worksheets[0]
    
    # 색상 정보 저장 (row, col) : hex_color
    cell_colors = {}
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                if cell.fill.fgColor.rgb not in ['00000000', None]:
                    cell_colors[(row-1, col-1)] = cell.fill.fgColor.rgb
    
    wb.close()
    
    print(f"색상이 있는 셀: {len(cell_colors)}개")
    
    # 4. 데이터를 Google Sheets 형식으로 변환
    # NaN 값을 빈 문자열로 변환
    df = df.fillna('')
    
    # 헤더 포함하여 데이터 준비
    values = [df.columns.tolist()] + df.values.tolist()
    
    # 모든 값을 문자열로 변환
    for i in range(len(values)):
        for j in range(len(values[i])):
            values[i][j] = str(values[i][j])
    
    print(f"총 {len(values)}개 행 (헤더 포함) 준비 완료")
    
    # 5. 시트 클리어
    print(f"{sheet_name} 시트 클리어 중...")
    clear_range = f"{sheet_name}!A:ZZ"
    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range=clear_range
    ).execute()
    
    # 6. 데이터 쓰기
    print(f"{sheet_name} 시트에 데이터 쓰는 중...")
    body = {
        'values': values
    }
    
    result = service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"{sheet_name}!A1",
        valueInputOption='RAW',
        body=body
    ).execute()
    
    # 7. 색상 적용
    if cell_colors:
        print(f"색상 적용 중... ({len(cell_colors)}개 셀)")
        
        requests = []
        sheet_id = get_sheet_id(service, spreadsheet_id, sheet_name)
        
        for (row, col), hex_color in cell_colors.items():
            rgb = hex_to_rgb(hex_color)
            if rgb:
                requests.append({
                    "updateCells": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": row,
                            "endRowIndex": row + 1,
                            "startColumnIndex": col,
                            "endColumnIndex": col + 1
                        },
                        "rows": [{
                            "values": [{
                                "userEnteredFormat": {
                                    "backgroundColor": rgb
                                }
                            }]
                        }],
                        "fields": "userEnteredFormat.backgroundColor"
                    }
                })
        
        if requests:
            batch_update_body = {"requests": requests}
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body=batch_update_body
            ).execute()
            print("✅ 색상 적용 완료")
    
    print(f"✅ 업데이트 완료: {result.get('updatedCells')}개 셀 업데이트됨")
    print(f"✅ 업데이트 범위: {result.get('updatedRange')}")

def get_sheet_id(service, spreadsheet_id, sheet_name):
    """시트 이름으로 시트 ID 가져오기"""
    spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheets = spreadsheet.get('sheets', [])
    
    for sheet in sheets:
        if sheet['properties']['title'] == sheet_name:
            return sheet['properties']['sheetId']
    
    raise ValueError(f"시트 '{sheet_name}'을 찾을 수 없습니다.")

def main():
    """계산된 OCR 결과를 Google Sheets에 업로드"""
    
    # Google Sheets API 인증
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    
    # 서비스 계정 키 파일 경로
    from pathlib import Path
    project_root = Path(__file__).parent.parent.parent
    SERVICE_ACCOUNT_FILE = project_root / "src" / "config" / "google_api_key.json"
    
    credentials = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    
    service = build('sheets', 'v4', credentials=credentials)
    
    # 스프레드시트 ID
    SPREADSHEET_ID = '1njdeOI4TLyF2IkggosBUGgg5yKetez8cdcepbsAeEx4'
    
    # OCR 결과 디렉토리
    ocr_result_dir = Path(__file__).parent / "OCR" / "OCR_결과" / "20250708"
    
    # 계산된 파일 찾기
    calculated_files = glob.glob(str(ocr_result_dir / '*_calculated.xlsx'))
    
    print("=" * 50)
    print("계산된 OCR 결과를 Google Sheets에 업로드")
    print("=" * 50)
    
    for file_path in calculated_files:
        file_name = os.path.basename(file_path)
        
        # 대리점명 추출
        if 'KT_' in file_name:
            sheet_name = 'kt_ocr_calculated'
            carrier = 'KT'
        elif 'LG_' in file_name:
            sheet_name = 'lg_ocr_calculated'
            carrier = 'LG'
        elif 'SK_' in file_name:
            sheet_name = 'sk_ocr_calculated'
            carrier = 'SK'
        else:
            print(f"알 수 없는 파일: {file_name}")
            continue
        
        print(f"\n{carrier} 파일 업데이트: {file_name}")
        
        try:
            update_google_sheet_with_colors(file_path, SPREADSHEET_ID, sheet_name, service)
        except Exception as e:
            print(f"❌ 오류 발생: {e}")
            print(f"시트 '{sheet_name}'가 존재하지 않을 수 있습니다. 시트를 생성하거나 이름을 확인하세요.")

if __name__ == "__main__":
    main()