import yaml
import openpyxl
from openpyxl.styles import Font, PatternFill
import re
import os
from datetime import datetime
from rich.console import Console
import unicodedata

console = Console()


class PolicyBasedCalculator:
    """정책 기반 단가 계산기"""
    
    def __init__(self, policy_file="policy_config.yaml"):
        """정책 파일 로드"""
        with open(policy_file, 'r', encoding='utf-8') as f:
            self.policies = yaml.safe_load(f)
        self.calculations_log = []
    
    def get_dealer_from_filename(self, filename):
        """파일명에서 대리점명 추출"""
        filename = unicodedata.normalize('NFC', filename)
        
        # 정책에 있는 대리점명과 매칭
        for dealer in self.policies.keys():
            if dealer.replace("_", "_") in filename:
                return dealer
        return None
    
    def is_pure_number(self, value):
        """순수 숫자인지 확인"""
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return True
        if isinstance(value, str):
            value_str = value.strip()
            # 010은 신규가입을 의미하므로 제외
            if value_str == "010":
                return False
            return bool(re.match(r'^-?\d+\.?\d*$', value_str))
        return False
    
    def extract_number(self, text):
        """텍스트에서 숫자 추출"""
        if not text:
            return 0
        text = str(text).strip()
        if re.match(r'^-?\d+\.?\d*$', text):
            return float(text)
        return 0
    
    def get_cell_background_color(self, cell):
        """셀의 배경색 가져오기"""
        if hasattr(cell, 'fill') and cell.fill and hasattr(cell.fill, 'start_color'):
            color = cell.fill.start_color
            if hasattr(color, 'rgb') and color.rgb:
                if isinstance(color.rgb, str) and len(color.rgb) >= 6:
                    rgb = color.rgb[-6:]
                    r = int(rgb[0:2], 16)
                    g = int(rgb[2:4], 16)
                    b = int(rgb[4:6], 16)
                    
                    if r > 200 and g > 200 and b < 100:
                        return "yellow"
                    elif b > 200 and r < 100 and g < 100:
                        return "blue"
                    elif r > 200 and g < 100 and b < 100:
                        return "red"
        return "none"
    
    def is_red_text(self, cell):
        """셀의 텍스트가 빨간색인지 확인"""
        if hasattr(cell, 'font') and cell.font and hasattr(cell.font, 'color'):
            color = cell.font.color
            if hasattr(color, 'rgb') and color.rgb:
                if isinstance(color.rgb, str) and len(color.rgb) >= 6:
                    rgb = color.rgb[-6:]
                    r = int(rgb[0:2], 16)
                    g = int(rgb[2:4], 16)
                    b = int(rgb[4:6], 16)
                    if r > 200 and g < 100 and b < 100:
                        return True
        return False
    
    def apply_rules(self, rules, value, context=None):
        """규칙 적용"""
        result = value
        
        for rule in rules:
            if 'condition' in rule:
                # 조건 평가 (간단한 구현)
                if rule['condition'] == 'default':
                    pass
                elif 'bg_color' in rule['condition'] and context:
                    if context.get('bg_color') not in rule['condition']:
                        continue
                elif 'text_color' in rule['condition'] and context:
                    if context.get('text_color') not in rule['condition']:
                        continue
                elif 'result > 30' in rule['condition']:
                    if result <= 30:
                        continue
                # 추가 조건들은 필요에 따라 구현
            
            # 수식 적용
            formula = rule['formula']
            if 'value' in formula:
                formula = formula.replace('value', str(result))
            if 'result' in formula:
                formula = formula.replace('result', str(result))
            
            try:
                # 간단한 수식 평가
                if '+' in formula:
                    parts = formula.split('+')
                    result = float(parts[0]) + float(parts[1])
                elif '-' in formula:
                    parts = formula.split('-')
                    result = float(parts[0]) - float(parts[1])
                elif '*' in formula:
                    parts = formula.split('*')
                    result = float(parts[0]) * float(parts[1])
                else:
                    result = float(formula)
                
                if 'name' in rule:
                    self.calculations_log.append(f"{rule['name']}: {formula} = {result}")
            except:
                pass
        
        return result
    
    def process_excel_file(self, file_path):
        """Excel 파일 처리"""
        try:
            filename = os.path.basename(file_path)
            dealer = self.get_dealer_from_filename(filename)
            
            if not dealer:
                console.print(f"[yellow]⚠️ {filename}: 대리점 정보를 찾을 수 없습니다[/yellow]")
                return None
            
            if dealer not in self.policies:
                console.print(f"[yellow]⚠️ {dealer}: 정책이 정의되지 않았습니다[/yellow]")
                return None
            
            policy = self.policies[dealer]
            console.print(f"\n[green]📊 {filename} 처리 중... (대리점: {dealer})[/green]")
            console.print(f"[blue]정책 버전: {policy['version']} - {policy['description']}[/blue]")
            
            # Excel 파일 열기
            wb = openpyxl.load_workbook(file_path)
            self.calculations_log = []
            
            # 각 시트 처리
            for sheet_name in wb.sheetnames:
                if "_분석" in sheet_name:
                    continue
                
                ws = wb[sheet_name]
                
                # 각 셀 처리
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and self.is_pure_number(cell.value):
                            original_value = self.extract_number(cell.value)
                            if original_value != 0:
                                # 컨텍스트 정보 수집
                                context = {
                                    'bg_color': self.get_cell_background_color(cell),
                                    'text_color': 'red' if self.is_red_text(cell) else 'black',
                                    'sheet': sheet_name,
                                    'cell': cell.coordinate
                                }
                                
                                # 규칙 적용
                                new_value = self.apply_rules(policy['rules'], original_value, context)
                                
                                # 최종 곱셈
                                final_value = new_value * policy['multiplier']
                                cell.value = final_value
                                
                                self.calculations_log.append(
                                    f"{sheet_name} {cell.coordinate}: {original_value} → {final_value}"
                                )
            
            # 결과 저장
            output_path = file_path.replace(".xlsx", "_calculated_v2.xlsx")
            wb.save(output_path)
            
            # 로그 출력
            if self.calculations_log:
                console.print(f"\n[cyan]계산 로그 (샘플 5개):[/cyan]")
                for log in self.calculations_log[:5]:
                    console.print(f"  {log}")
                if len(self.calculations_log) > 5:
                    console.print(f"  ... 외 {len(self.calculations_log) - 5}개의 계산")
            
            console.print(f"[green]✅ 계산 완료: {output_path}[/green]")
            return output_path
            
        except Exception as e:
            console.print(f"[red]❌ 처리 중 오류 발생: {str(e)}[/red]")
            return None


def main():
    """메인 함수"""
    calculator = PolicyBasedCalculator()
    
    # OCR 결과 폴더에서 파일 찾기
    from pathlib import Path
    ocr_result_dir = Path(__file__).parent / "OCR" / "OCR_결과" / "20250708"
    
    if not ocr_result_dir.exists():
        console.print(f"[red]❌ OCR 결과 폴더를 찾을 수 없습니다: {ocr_result_dir}[/red]")
        return
    
    # tables 파일 처리
    excel_files = []
    for file in ocr_result_dir.iterdir():
        if file.name.endswith("_tables.xlsx") and not file.name.endswith("_calculated.xlsx"):
            excel_files.append(str(file))
    
    if not excel_files:
        console.print("[red]❌ 처리할 Excel 파일을 찾을 수 없습니다.[/red]")
        return
    
    console.print(f"[green]🔍 {len(excel_files)}개의 Excel 파일을 찾았습니다.[/green]")
    
    # 각 파일 처리
    processed_files = []
    for file_path in excel_files:
        result = calculator.process_excel_file(file_path)
        if result:
            processed_files.append(result)
    
    console.print(f"\n[bold green]✨ 처리 완료![/bold green]")
    console.print(f"총 {len(processed_files)}개 파일 처리됨")


if __name__ == "__main__":
    main()