import pandas as pd
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import load_workbook
import time
import socket
from googleapiclient.errors import HttpError

def hex_to_rgb(hex_color):
    """16진수 색상을 RGB 딕셔너리로 변환"""
    if not hex_color or hex_color == '00000000':
        return None
    
    # ARGB 형식인 경우 앞의 2자리(Alpha) 제거
    if len(hex_color) == 8:
        hex_color = hex_color[2:]
    
    try:
        r = int(hex_color[0:2], 16) / 255.0
        g = int(hex_color[2:4], 16) / 255.0
        b = int(hex_color[4:6], 16) / 255.0
        return {"red": r, "green": g, "blue": b}
    except:
        return None

def update_google_sheet_with_colors(excel_file_path, spreadsheet_id, sheet_name):
    """엑셀 파일의 데이터와 색상을 Google Sheets에 업데이트"""
    
    # 1. 엑셀 파일 읽기
    print(f"엑셀 파일 읽는 중: {excel_file_path}")
    df = pd.read_excel(excel_file_path)
    
    # 2. 엑셀 파일에서 색상 정보 추출
    wb = load_workbook(excel_file_path, data_only=True)
    ws = wb.worksheets[0]
    
    # 색상 정보 저장 (row, col) : hex_color
    cell_colors = {}
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                if cell.fill.fgColor.rgb not in ['00000000', None]:
                    cell_colors[(row-1, col-1)] = cell.fill.fgColor.rgb
    
    wb.close()
    
    print(f"색상이 있는 셀: {len(cell_colors)}개")
    
    # 3. Google Sheets API 인증
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    
    try:
        # 인증 정보 생성
        # 외부 config 폴더의 Google API 키 사용
        from pathlib import Path
        key_file = Path("/Users/jacob_athometrip/Desktop/dev/nofee/workspace_nofee/config/google_api_key.json")
        
        credentials = service_account.Credentials.from_service_account_file(
            str(key_file), scopes=SCOPES)
        
        # Google Sheets API 서비스 생성
        service = build('sheets', 'v4', credentials=credentials)
        
        # 4. 데이터를 2차원 리스트로 변환
        headers = df.columns.tolist()
        # device_name을 device_nm으로 변경
        headers = ['device_nm' if col == 'device_name' else col for col in headers]
        values = [headers]  # 헤더
        
        for _, row in df.iterrows():
            row_values = [str(val) for val in row]
            values.append(row_values)
        
        print(f"총 {len(values)}개 행 (헤더 포함) 준비 완료")
        
        # 5. 시트 클리어
        print(f"{sheet_name} 시트 클리어 중...")
        service.spreadsheets().values().clear(
            spreadsheetId=spreadsheet_id,
            range=f"{sheet_name}!A:ZZ"
        ).execute()
        
        # 6. 새 데이터 쓰기
        print(f"{sheet_name} 시트에 데이터 쓰는 중...")
        body = {'values': values}
        
        result = service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"{sheet_name}!A1",
            valueInputOption='RAW',
            body=body
        ).execute()
        
        # 7. 색상 적용을 위한 배치 업데이트 요청 준비
        requests = []
        
        # 시트 ID 가져오기
        sheet_metadata = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheet_id = None
        for sheet in sheet_metadata.get('sheets', []):
            if sheet.get('properties', {}).get('title') == sheet_name:
                sheet_id = sheet.get('properties', {}).get('sheetId')
                break
        
        if sheet_id is not None:
            # 색상 적용 요청 생성
            for (row, col), hex_color in cell_colors.items():
                rgb = hex_to_rgb(hex_color)
                if rgb:
                    requests.append({
                        "updateCells": {
                            "range": {
                                "sheetId": sheet_id,
                                "startRowIndex": row,
                                "endRowIndex": row + 1,
                                "startColumnIndex": col,
                                "endColumnIndex": col + 1
                            },
                            "rows": [{
                                "values": [{
                                    "userEnteredFormat": {
                                        "backgroundColor": rgb
                                    }
                                }]
                            }],
                            "fields": "userEnteredFormat.backgroundColor"
                        }
                    })
            
            # 배치 업데이트 실행
            if requests:
                print(f"색상 적용 중... ({len(requests)}개 셀)")
                batch_update_body = {"requests": requests}
                service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body=batch_update_body
                ).execute()
                print("✅ 색상 적용 완료")
        
        print(f"✅ 업데이트 완료: {result.get('updatedCells')}개 셀 업데이트됨")
        print(f"✅ 업데이트 범위: {result.get('updatedRange')}")
        
    except FileNotFoundError:
        print("❌ 서비스 계정 키 파일을 찾을 수 없습니다.")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")

def update_kt_sheet_with_colors():
    """KT 데이터를 색상과 함께 Google Sheets에 업데이트"""
    from shared_config.config.paths import PathManager

    spreadsheet_id = '1njdeOI4TLyF2IkggosBUGgg5yKetez8cdcepbsAeEx4'
    sheet_name = 'kt_price'

    pm = PathManager()

    # rebated latest 파일 우선, 없으면 merged latest 파일 사용
    kt_rebated_file = pm.merged_latest_dir / 'kt_rebated_with_colors_latest.xlsx'
    kt_merged_file = pm.merged_latest_dir / 'kt_merged_with_colors_latest.xlsx'

    if kt_rebated_file.exists():
        print(f"KT 파일 업데이트 (rebated): {kt_rebated_file.name}")
        update_google_sheet_with_colors(str(kt_rebated_file), spreadsheet_id, sheet_name)
    elif kt_merged_file.exists():
        print(f"KT 파일 업데이트 (merged): {kt_merged_file.name}")
        update_google_sheet_with_colors(str(kt_merged_file), spreadsheet_id, sheet_name)
    else:
        print("⚠️ KT 파일을 찾을 수 없습니다.")

def update_sk_sheet_with_colors():
    """SK 데이터를 색상과 함께 Google Sheets에 업데이트"""
    from shared_config.config.paths import PathManager

    spreadsheet_id = '1njdeOI4TLyF2IkggosBUGgg5yKetez8cdcepbsAeEx4'
    sheet_name = 'sk_price'

    pm = PathManager()

    # rebated latest 파일 우선, 없으면 merged latest 파일 사용
    sk_rebated_file = pm.merged_latest_dir / 'sk_rebated_latest.xlsx'
    sk_merged_file = pm.merged_latest_dir / 'sk_merged_latest.xlsx'

    if sk_rebated_file.exists():
        print(f"SK 파일 업데이트 (rebated): {sk_rebated_file.name}")
        update_google_sheet_with_colors(str(sk_rebated_file), spreadsheet_id, sheet_name)
    elif sk_merged_file.exists():
        print(f"SK 파일 업데이트 (merged): {sk_merged_file.name}")
        update_google_sheet_with_colors(str(sk_merged_file), spreadsheet_id, sheet_name)
    else:
        print("⚠️ SK 파일을 찾을 수 없습니다.")

def update_lg_sheet_with_colors():
    """LG 데이터를 색상과 함께 Google Sheets에 업데이트"""
    from shared_config.config.paths import PathManager

    spreadsheet_id = '1njdeOI4TLyF2IkggosBUGgg5yKetez8cdcepbsAeEx4'
    sheet_name = 'lg_price'

    pm = PathManager()

    # rebated latest 파일 우선, 없으면 merged latest 파일 사용
    lg_rebated_file = pm.merged_latest_dir / 'lg_rebated_latest.xlsx'
    lg_merged_file = pm.merged_latest_dir / 'lg_merged_latest.xlsx'

    if lg_rebated_file.exists():
        print(f"LG 파일 업데이트 (rebated): {lg_rebated_file.name}")
        update_google_sheet_with_colors(str(lg_rebated_file), spreadsheet_id, sheet_name)
    elif lg_merged_file.exists():
        print(f"LG 파일 업데이트 (merged): {lg_merged_file.name}")
        update_google_sheet_with_colors(str(lg_merged_file), spreadsheet_id, sheet_name)
    else:
        print("⚠️ LG 파일을 찾을 수 없습니다.")

if __name__ == "__main__":
    print("색상을 유지하여 Google Sheets 업데이트")
    print("="*50)
    
    # KT 데이터 업데이트
    update_kt_sheet_with_colors()
    
    # 다른 통신사도 필요시 추가
    update_sk_sheet_with_colors()
    update_lg_sheet_with_colors()