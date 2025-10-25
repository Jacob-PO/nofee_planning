import pandas as pd
import os
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import unicodedata

def extract_data_from_kt_dableu(file_path, date, carrier, dealer):
    """KT 더블유 파일에서 데이터 추출"""
    df = pd.read_excel(file_path, sheet_name=0)
    
    # 원본 워크북 열기 (색상 정보용)
    wb = load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]
    
    device_data = {}
    color_info = {}  # 색상이 있는 셀 정보만 저장
    
    plan_mapping = [
        {'cols': [1, 2, 3], 'plan': '110k', 'types': ['신규가입', '번호이동', '기기변경']},
        {'cols': [4, 5, 6], 'plan': '100k', 'types': ['신규가입', '번호이동', '기기변경']},
        {'cols': [7, 8, 9], 'plan': '90k', 'types': ['신규가입', '번호이동', '기기변경']},
        {'cols': [10, 11, 12], 'plan': '61k', 'types': ['신규가입', '번호이동', '기기변경']},
        {'cols': [13, 14, 15], 'plan': '37k', 'types': ['신규가입', '번호이동', '기기변경']},
    ]
    
    for idx in range(3, len(df)):
        device_name = df.iloc[idx, 0]
        
        if pd.isna(device_name) or device_name == '':
            continue
            
        # 헤더 텍스트 건너뛰기
        if device_name in ['모델', '모델명', 'device', 'Device']:
            continue
        
        if device_name not in device_data:
            device_data[device_name] = {
                'date': date,
                'carrier': carrier,
                'dealer': dealer,
                'device_name': device_name
            }
            color_info[device_name] = {}
        
        # 기기명 셀의 색상 확인 (색상이 있는 경우만 저장)
        cell = ws.cell(row=idx+1, column=1)
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            if cell.fill.fgColor.rgb not in ['00000000', None]:  # 투명이나 None이 아닌 경우만
                color_info[device_name]['device_color'] = cell.fill.fgColor.rgb
        
        for plan_info in plan_mapping:
            plan = plan_info['plan']
            for col_offset, join_type in enumerate(plan_info['types']):
                col_idx = plan_info['cols'][col_offset]
                if col_idx < len(df.columns):
                    value = df.iloc[idx, col_idx]
                    key = f"{join_type}_공시_{plan}"
                    device_data[device_name][key] = value if pd.notna(value) else None
                    # KT도 선택약정을 동일한 값으로 추가 (SK, LG와 동일한 방식)
                    key_선약 = f"{join_type}_선약_{plan}"
                    device_data[device_name][key_선약] = value if pd.notna(value) else None
                    
                    # 데이터 셀의 색상 확인 (색상이 있는 경우만 저장)
                    cell = ws.cell(row=idx+1, column=col_idx+1)
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                        if cell.fill.fgColor.rgb not in ['00000000', None]:
                            color_key = f'{key}_color'
                            color_info[device_name][color_key] = cell.fill.fgColor.rgb
                            # 선약 셀에도 동일한 색상 적용
                            color_key_선약 = f'{key_선약}_color'
                            color_info[device_name][color_key_선약] = cell.fill.fgColor.rgb
    
    wb.close()
    return list(device_data.values()), color_info

def extract_data_from_번개폰(file_path, date, carrier, dealer):
    """번개폰 파일에서 데이터 추출 (KT용)"""
    df = pd.read_excel(file_path, sheet_name=0)

    device_data = {}

    # 번개폰은 2행부터 데이터 시작 (0행: 헤더, 1행: 컬럼명)
    for idx in range(2, len(df)):
        device_name = df.iloc[idx, 0]

        if pd.isna(device_name) or device_name == '':
            continue

        if device_name not in device_data:
            device_data[device_name] = {
                'date': date,
                'carrier': carrier,
                'dealer': dealer,
                'device_name': device_name
            }

        # KT 번개폰 실제 컬럼 구조:
        # 0열: 기기명
        # 1열: 110k 번호이동 선약
        # 2열: 110k 기기변경 선약
        # 3열: 109k 번호이동 공시
        # 4열: 109k 기기변경 공시

        # 1열: 110k 번호이동 선약
        if pd.notna(df.iloc[idx, 1]):
            device_data[device_name]['번호이동_선약_110k'] = df.iloc[idx, 1]
            device_data[device_name]['번호이동_공시_110k'] = df.iloc[idx, 1]  # KT는 공시/선약 동일

        # 2열: 110k 기기변경 선약
        if pd.notna(df.iloc[idx, 2]):
            device_data[device_name]['기기변경_선약_110k'] = df.iloc[idx, 2]
            device_data[device_name]['기기변경_공시_110k'] = df.iloc[idx, 2]  # KT는 공시/선약 동일

        # 3열: 109k 번호이동 공시 (100k로 매핑)
        if pd.notna(df.iloc[idx, 3]):
            device_data[device_name]['번호이동_공시_100k'] = df.iloc[idx, 3]
            device_data[device_name]['번호이동_선약_100k'] = df.iloc[idx, 3]

        # 4열: 109k 기기변경 공시 (100k로 매핑)
        if pd.notna(df.iloc[idx, 4]):
            device_data[device_name]['기기변경_공시_100k'] = df.iloc[idx, 4]
            device_data[device_name]['기기변경_선약_100k'] = df.iloc[idx, 4]

    return list(device_data.values()), {}  # 번개폰은 색상 정보 없음

def extract_data_from_kt_max(file_path, date, carrier, dealer):
    """KT 맥스 파일에서 데이터 추출"""
    df = pd.read_excel(file_path, sheet_name=0)
    
    # 원본 워크북 열기 (색상 정보용)
    wb = load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]
    
    device_data = {}
    color_info = {}  # 색상이 있는 셀 정보만 저장
    
    plan_mapping = [
        {'cols': [1, 2, 3], 'plan': '110k', 'types': ['신규가입', '번호이동', '기기변경']},
        {'cols': [4, 5, 6], 'plan': '100k', 'types': ['신규가입', '번호이동', '기기변경']},
        {'cols': [7, 8, 9], 'plan': '90k', 'types': ['신규가입', '번호이동', '기기변경']},
        {'cols': [10, 11, 12], 'plan': '61k', 'types': ['신규가입', '번호이동', '기기변경']},
        {'cols': [13, 14, 15], 'plan': '37k', 'types': ['신규가입', '번호이동', '기기변경']},
    ]
    
    for idx in range(4, len(df)):  # KT 맥스는 4행부터 시작
        device_name = df.iloc[idx, 0]
        
        if pd.isna(device_name) or device_name == '':
            continue
            
        # 헤더 텍스트 건너뛰기
        if device_name in ['모델', '모델명', 'device', 'Device']:
            continue
        
        if device_name not in device_data:
            device_data[device_name] = {
                'date': date,
                'carrier': carrier,
                'dealer': dealer,
                'device_name': device_name
            }
            color_info[device_name] = {}
        
        # 기기명 셀의 색상 확인 (색상이 있는 경우만 저장)
        cell = ws.cell(row=idx+1, column=1)
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            if cell.fill.fgColor.rgb not in ['00000000', None]:
                color_info[device_name]['device_color'] = cell.fill.fgColor.rgb
        
        for plan_info in plan_mapping:
            plan = plan_info['plan']
            for col_offset, join_type in enumerate(plan_info['types']):
                col_idx = plan_info['cols'][col_offset]
                if col_idx < len(df.columns):
                    value = df.iloc[idx, col_idx]
                    key = f"{join_type}_공시_{plan}"
                    device_data[device_name][key] = value if pd.notna(value) else None
                    # KT도 선택약정을 동일한 값으로 추가 (SK, LG와 동일한 방식)
                    key_선약 = f"{join_type}_선약_{plan}"
                    device_data[device_name][key_선약] = value if pd.notna(value) else None
                    
                    # 데이터 셀의 색상 확인 (색상이 있는 경우만 저장)
                    cell = ws.cell(row=idx+1, column=col_idx+1)
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                        if cell.fill.fgColor.rgb not in ['00000000', None]:
                            color_key = f'{key}_color'
                            color_info[device_name][color_key] = cell.fill.fgColor.rgb
                            # 선약 셀에도 동일한 색상 적용
                            color_key_선약 = f'{key_선약}_color'
                            color_info[device_name][color_key_선약] = cell.fill.fgColor.rgb
    
    wb.close()
    return list(device_data.values()), color_info

# KT 사전예약 관련 함수 제거

def merge_kt_files_with_colors():
    """모든 KT 파일을 색상과 함께 병합"""
    from shared_config.config.data_merge_config import find_latest_ocr_results_folder
    
    base_path = find_latest_ocr_results_folder()
    if not base_path:
        return None
    
    # 폴더명에서 날짜 추출 (20250714 -> 2025. 7. 14)
    folder_name = os.path.basename(base_path)
    if len(folder_name) == 8 and folder_name.isdigit():
        year = folder_name[:4]
        month = str(int(folder_name[4:6]))  # 앞의 0 제거
        day = folder_name[6:8]              # 일자는 2자리 유지
        date_str = f"{year}. {month}. {day}"
    else:
        date_str = "2025. 7. 14"  # 기본값
    
    all_data = []
    all_color_info = {}
    
    # KT 더블유 파일 처리 - calculated 파일만 사용
    dableu_file = None
    for file in os.listdir(base_path):
        normalized_file = unicodedata.normalize('NFC', file)
        if "KT" in normalized_file and "더블유" in normalized_file and file.endswith("_calculated.xlsx"):
            dableu_file = os.path.join(base_path, file)
            break
    if dableu_file and os.path.exists(dableu_file):
        print(f"KT 더블유 파일 처리 중: {os.path.basename(dableu_file)}")
        dableu_data, dableu_colors = extract_data_from_kt_dableu(dableu_file, date_str, "KT", "더블유")
        all_data.extend(dableu_data)
        all_color_info.update(dableu_colors)
        print(f"KT 더블유: {len(dableu_data)}개 기기 추출")
        
        # 색상 정보 확인
        color_count = sum(len(colors) for colors in dableu_colors.values())
        print(f"  색상이 있는 셀: {color_count}개")
    
    # KT 맥스 파일 처리 - calculated 파일만 사용
    max_file = None
    for file in os.listdir(base_path):
        normalized_file = unicodedata.normalize('NFC', file)
        if "KT" in normalized_file and "맥스" in normalized_file and file.endswith("_calculated.xlsx"):
            max_file = os.path.join(base_path, file)
            break
    if max_file and os.path.exists(max_file):
        print(f"KT 맥스 파일 처리 중: {os.path.basename(max_file)}")
        max_data, max_colors = extract_data_from_kt_max(max_file, date_str, "KT", "맥스")
        all_data.extend(max_data)
        all_color_info.update(max_colors)
        print(f"KT 맥스: {len(max_data)}개 기기 추출")
        
        # 색상 정보 확인
        color_count = sum(len(colors) for colors in max_colors.values())
        print(f"  색상이 있는 셀: {color_count}개")
    
    # KT 번개폰 파일 처리 - calculated 파일만 사용
    bungyae_file = None
    for file in os.listdir(base_path):
        normalized_file = unicodedata.normalize('NFC', file)
        if "KT" in normalized_file and "번개폰" in normalized_file and file.endswith("_calculated.xlsx"):
            bungyae_file = os.path.join(base_path, file)
            break
    if bungyae_file and os.path.exists(bungyae_file):
        print(f"KT 번개폰 파일 처리 중: {os.path.basename(bungyae_file)}")
        bungyae_data, bungyae_colors = extract_data_from_번개폰(bungyae_file, date_str, "KT", "번개폰")
        all_data.extend(bungyae_data)
        all_color_info.update(bungyae_colors)
        print(f"KT 번개폰: {len(bungyae_data)}개 기기 추출")

    # 애플 사전예약 데이터 추가
    try:
        from data_merge.apple_preorder_merge import process_apple_preorder_data
        apple_df = process_apple_preorder_data('KT')
        if apple_df is not None and not apple_df.empty:
            # DataFrame을 딕셔너리 리스트로 변환
            apple_data = apple_df.to_dict('records')
            all_data.extend(apple_data)
            print(f"KT 애플사전예약: {len(apple_data)}개 기기 추출")
    except Exception as e:
        print(f"애플 사전예약 데이터 처리 중 오류: {e}")
    
    if all_data:
        df = pd.DataFrame(all_data)
        
        # 모든 가능한 컬럼 정의
        base_columns = ['date', 'carrier', 'dealer', 'device_name']
        plan_types = ['110k', '100k', '90k', '61k', '37k']
        join_types = ['신규가입', '번호이동', '기기변경']
        support_types = ['공시', '선약']
        
        dynamic_columns = []
        for join_type in join_types:
            for support_type in support_types:
                for plan_type in plan_types:
                    dynamic_columns.append(f'{join_type}_{support_type}_{plan_type}')
        
        all_columns = base_columns + dynamic_columns
        
        for col in all_columns:
            if col not in df.columns:
                df[col] = None
        
        df = df[all_columns]
        df = df.fillna('Null')
        
        # PathManager를 사용한 출력 경로 설정
        from shared_config.config.paths import PathManager
        pm = PathManager()
        archive_path, latest_path = pm.get_merged_output_path('kt', is_rebated=False, with_colors=True)

        # 결과 저장 (색상 포함)
        output_file = archive_path
        
        with pd.ExcelWriter(str(output_file), engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='KT_price')
            
            # 워크시트 가져오기
            worksheet = writer.sheets['KT_price']
            
            # 색상 적용 (색상이 있는 셀만)
            for row_idx, row in df.iterrows():
                device_name = row['device_name']
                
                if device_name in all_color_info and all_color_info[device_name]:
                    color_dict = all_color_info[device_name]
                    
                    # 기기명 셀에 색상 적용
                    if 'device_color' in color_dict:
                        cell = worksheet.cell(row=row_idx+2, column=4)  # device_name 열
                        cell.fill = PatternFill(start_color=color_dict['device_color'], 
                                              end_color=color_dict['device_color'], 
                                              fill_type="solid")
                    
                    # 각 데이터 셀에 색상 적용
                    for col_idx, col_name in enumerate(df.columns):
                        color_key = f'{col_name}_color'
                        if color_key in color_dict:
                            cell = worksheet.cell(row=row_idx+2, column=col_idx+1)
                            cell.fill = PatternFill(start_color=color_dict[color_key], 
                                                  end_color=color_dict[color_key], 
                                                  fill_type="solid")
        
        # latest 파일로 복사
        pm.save_with_archive(archive_path, archive_path, latest_path)

        print(f"\n병합 완료!")
        print(f"총 {len(df)}개 기기")
        print(f"📁 Archive: {archive_path}")
        print(f"📁 Latest: {latest_path}")
        print("원본 파일의 색상이 있는 셀만 색상이 적용되었습니다.")

        return df
    else:
        print("추출된 데이터가 없습니다.")
        return None

if __name__ == "__main__":
    df = merge_kt_files_with_colors()